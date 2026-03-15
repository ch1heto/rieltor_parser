import argparse
import asyncio
import random
import re
import time
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from urllib.parse import urlparse, parse_qs

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from playwright_stealth import Stealth as _Stealth

    async def _apply_stealth(page):
        await _Stealth().apply_stealth_async(page)

    STEALTH_AVAILABLE = True
except ImportError:
    try:
        from playwright_stealth import stealth_async as _apply_stealth  # type: ignore
        STEALTH_AVAILABLE = True
    except ImportError:
        async def _apply_stealth(page):
            return None
        STEALTH_AVAILABLE = False

OUTPUT_DIR = Path(__file__).parent
HEADLESS = False
PARALLEL_TABS = 10
CATALOG_BROWSERS = 1  # параллельных браузера для сбора каталога
RETRY_COUNT = 3
AUTOSAVE_EVERY = 20
CATALOG_DELAY = (0.5, 1.0)
PROFILE_DELAY = (0.8, 1.6)
REGION_PAUSE = (6.0, 12.0)
MAX_PAGES = 0
MAX_AGENTS = 0

BASE_DOMAIN_DEFAULT = "https://cian.ru"
CATALOG_DOMAIN = "https://chelyabinsk.cian.ru"
WARMUP_REGION_ID = 4630

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(OUTPUT_DIR / "cian_parser.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)


async def new_stealth_page(ctx):
    page = await ctx.new_page()
    await _apply_stealth(page)
    return page


def sanitize_tag(s: str) -> str:
    s = re.sub(r"[^\w\-]+", "_", s.strip(), flags=re.U)
    s = re.sub(r"_+", "_", s)
    return s.strip("_") or "region"


def normalize_ws(s: str) -> str:
    s = (s or "").replace("\xa0", " ").replace("\u2009", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def build_catalog_url(region_id: int, page_num: int = 1, base_domain: str = CATALOG_DOMAIN) -> str:
    return f"{base_domain.rstrip('/')}/realtors/?regionId={region_id}&page={page_num}"


def parse_region_id_from_url(url: str) -> int:
    try:
        qs = parse_qs(urlparse(url).query)
        raw = qs.get("regionId", ["0"])[0]
        return int(raw)
    except Exception:
        return 0


def clean_phone(raw: str) -> str:
    raw = normalize_ws(raw or "")
    if not raw:
        return ""

    if "X" in raw.upper():
        m = re.search(r"(\+?7[\d\-\sXx]{8,})", raw)
        if m:
            val = normalize_ws(m.group(1))
            if not val.startswith("+"):
                val = "+" + val
            return val
        return raw

    digits = re.sub(r"\D", "", raw)
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    if len(digits) == 10:
        digits = "7" + digits
    if len(digits) >= 11:
        return f"+{digits[0]}-{digits[1:4]}-{digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
    return raw


def extract_email_from_text(t: str) -> str:
    t = t or ""
    m = re.search(r"([A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,})", t, re.I)
    return m.group(1) if m else ""


def extract_region_from_text(t: str) -> str:
    t = t or ""
    m = re.search(r"(Регион работы|Регионы работы|География работы)\s*[:\-]?\s*([^\n]{1,160})", t, re.I)
    if m:
        return normalize_ws(m.group(2))
    return ""


def extract_work_exp_from_text(t: str) -> str:
    t = t or ""
    m = re.search(r"(Опыт работы|Стаж)\s*[:\-]?\s*([^\n]{1,80})", t, re.I)
    if m:
        return normalize_ws(m.group(2))
    m = re.search(r"\bс\s*(\d{4})\s*г", t, re.I)
    if m:
        return f"с {m.group(1)} года"
    return ""


def extract_cian_from_text(t: str) -> str:
    t = t or ""
    m = re.search(r"(\d+\s*(?:лет|год|года)\s*(?:\s*\d+\s*(?:мес|месяц|месяца|месяцев))?\s*(?:на\s*Циан)?)", t, re.I)
    if m:
        return normalize_ws(m.group(1).replace("на Циан", "").strip())
    m = re.search(r"(\d+\s*(?:мес|месяц|месяца|месяцев))", t, re.I)
    if m:
        return normalize_ws(m.group(1))
    return ""


def extract_objects_from_text(t: str) -> str:
    t = t or ""
    m = re.search(r"(\d+)\s*(объект|объекта|объектов)\b", t, re.I)
    return normalize_ws(m.group(0)) if m else ""


def agents_url_from_href(href: str) -> str:
    if not href:
        return ""
    href = re.sub(r"[?#].*$", "", href.strip())
    m = re.search(r"/agents/(\d+)", href)
    if not m:
        return ""
    return f"{BASE_DOMAIN_DEFAULT}/agents/{m.group(1)}/"


async def make_browser_context(pw, headless: bool):
    browser = await pw.chromium.launch(
        headless=headless,
        args=[
            "--no-sandbox",
            "--disable-blink-features=AutomationControlled",
            "--disable-infobars",
            "--disable-dev-shm-usage",
            "--disable-extensions",
        ],
    )
    ctx = await browser.new_context(
        viewport={"width": 1440, "height": 900},
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        locale="ru-RU",
        timezone_id="Europe/Moscow",
        extra_http_headers={
            "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "sec-ch-ua": '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
        },
    )
    await ctx.add_init_script(
        "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
        "window.chrome={runtime:{}};"
        "Object.defineProperty(navigator,'plugins',{get:()=>[1,2,3,4,5]});"
        "Object.defineProperty(navigator,'languages',{get:()=>['ru-RU','ru','en-US','en']});"
    )

    BLOCKED_DOMAINS = (
        "google-analytics.com", "googletagmanager.com", "mc.yandex.ru",
        "top-fwz1.mail.ru", "counter.yadro.ru", "hotjar.com",
        "doubleclick.net", "facebook.net", "vk.com/rtrg",
        "sentry.io", "datadome.co", "adfox.ru",
    )

    async def route_handler(route):
        try:
            rt = route.request.resource_type
            url = route.request.url
            if rt in ("image", "media", "font", "stylesheet"):
                await route.abort()
                return
            if any(d in url for d in BLOCKED_DOMAINS):
                await route.abort()
                return
        except Exception:
            pass
        await route.continue_()

    await ctx.route("**/*", route_handler)
    return browser, ctx


async def safe_text(el) -> str:
    try:
        return (await el.inner_text()).strip()
    except Exception:
        return ""


async def text_or_empty(locator) -> str:
    try:
        if await locator.count() == 0:
            return ""
        return normalize_ws(await locator.first.inner_text())
    except Exception:
        return ""


async def has_catalog_cards(page, timeout: int = 3000) -> bool:
    try:
        await page.wait_for_selector('div[data-name="RealtorCard"], div[data-name="AgentCard"]', timeout=timeout)
        return True
    except Exception:
        return False


async def click_any(page, selectors: List[str], timeout: int = 4500) -> bool:
    for sel in selectors:
        try:
            locs = page.locator(sel)
            count = await locs.count()
            for i in range(min(count, 6)):
                loc = locs.nth(i)
                try:
                    await loc.scroll_into_view_if_needed(timeout=timeout)
                except Exception:
                    pass
                try:
                    await page.wait_for_timeout(random.uniform(300, 700))
                    await loc.click(timeout=timeout)
                    await page.wait_for_timeout(random.uniform(1200, 2200))
                    return True
                except Exception:
                    try:
                        await loc.click(timeout=timeout, force=True)
                        await page.wait_for_timeout(random.uniform(1200, 2200))
                        return True
                    except Exception:
                        continue
        except Exception:
            continue
    return False


async def ensure_catalog_open(page) -> bool:
    warm_url = build_catalog_url(WARMUP_REGION_ID, 1, CATALOG_DOMAIN)
    for _ in range(RETRY_COUNT):
        try:
            await page.goto(warm_url, wait_until="domcontentloaded", timeout=45_000)
            await page.wait_for_timeout(random.uniform(1200, 2200))
            if await has_catalog_cards(page, 12000):
                return True

            body = ""
            try:
                body = await page.evaluate("document.body ? document.body.innerText : ''")
            except Exception:
                body = ""

            if any(x in (body or "").lower() for x in ("подберём риелтора", "подберем риэлтора", "выберите риелтора", "под любую задачу")):
                clicked = await click_any(page, [
                    "button:has-text('Выбрать риелтора')",
                    "a:has-text('Выбрать риелтора')",
                    "button:has-text('Подобрать риелтора')",
                    "a:has-text('Подобрать риелтора')",
                    "button:has-text('Найти риелтора')",
                    "a:has-text('Найти риелтора')",
                    "a[href*='/realtors/']",
                ])
                if clicked and await has_catalog_cards(page, 15000):
                    return True

                await page.goto(warm_url, wait_until="domcontentloaded", timeout=45_000)
                await page.wait_for_timeout(random.uniform(1200, 2200))
                if await has_catalog_cards(page, 15000):
                    return True
        except Exception:
            await page.wait_for_timeout(random.uniform(2000, 4000))
    return False


async def open_region_modal(page) -> bool:
    btn = page.locator("button[data-name='GeoLocationButton']").first
    if await btn.count() == 0:
        return False
    try:
        await btn.scroll_into_view_if_needed(timeout=3000)
    except Exception:
        pass
    try:
        await btn.click(timeout=5000)
    except Exception:
        await btn.click(timeout=5000, force=True)
    try:
        await page.wait_for_selector("text=Выберите регион", timeout=8000)
        return True
    except Exception:
        return False


async def get_all_region_names(page) -> List[str]:
    ok = await open_region_modal(page)
    if not ok:
        return []

    names: List[str] = []

    try:
        loc = page.locator("div[data-name='PopularRegionsDesktop'] p")
        n = await loc.count()
        for i in range(n):
            t = normalize_ws(await loc.nth(i).inner_text())
            if t and t not in names:
                names.append(t)
    except Exception:
        pass

    try:
        loc = page.locator("div[data-name='SpecialRegionsDesktop'] label span")
        n = await loc.count()
        for i in range(n):
            t = normalize_ws(await loc.nth(i).inner_text())
            if t and t not in names:
                names.insert(0, t)
    except Exception:
        pass

    try:
        close_btn = page.locator("text=Выберите регион").locator("xpath=ancestor::*[1]").locator("xpath=following::*[name()='svg'][1]").first
        if await close_btn.count():
            try:
                await close_btn.click(timeout=2000)
            except Exception:
                pass
    except Exception:
        pass

    return names


async def get_selected_region_name(page) -> str:
    return await text_or_empty(page.locator("button[data-name='GeoLocationButton'] span"))


async def select_region(page, region_name: str) -> Tuple[bool, int]:
    current_region_id = parse_region_id_from_url(page.url or "")
    current_region_name = await get_selected_region_name(page)

    if current_region_name == region_name and current_region_id:
        return True, current_region_id

    ok = await open_region_modal(page)
    if not ok:
        return False, 0

    clicked = False

    try:
        target = page.locator("div[data-name='SpecialRegionsDesktop']").get_by_text(region_name, exact=True).first
        if await target.count():
            try:
                await target.scroll_into_view_if_needed(timeout=3000)
            except Exception:
                pass
            try:
                await target.click(timeout=5000)
            except Exception:
                await target.click(timeout=5000, force=True)
            clicked = True
    except Exception:
        pass

    if not clicked:
        try:
            target = page.locator("div[data-name='PopularRegionsDesktop']").get_by_text(region_name, exact=True).first
            if await target.count():
                try:
                    await target.scroll_into_view_if_needed(timeout=5000)
                except Exception:
                    pass
                try:
                    await target.click(timeout=5000)
                except Exception:
                    await target.click(timeout=5000, force=True)
                clicked = True
        except Exception:
            pass

    if not clicked:
        return False, 0

    for _ in range(30):
        await page.wait_for_timeout(500)
        rid = parse_region_id_from_url(page.url or "")
        selected_name = await get_selected_region_name(page)
        if selected_name == region_name and rid and (rid != current_region_id or region_name == current_region_name):
            if await has_catalog_cards(page, 12000):
                return True, rid

    rid = parse_region_id_from_url(page.url or "")
    if rid and await has_catalog_cards(page, 12000):
        return True, rid

    return False, 0


async def get_last_page(page) -> int:
    try:
        items = page.locator("div[data-name='PaginationWrapper'] [data-name='PaginationItem'] span")
        n = await items.count()
        nums = []
        for i in range(n):
            t = normalize_ws(await items.nth(i).inner_text())
            if t.isdigit():
                nums.append(int(t))
        return max(nums) if nums else 1
    except Exception:
        return 1


async def extract_catalog_cards(page) -> List[Dict]:
    cards = page.locator('div[data-name="RealtorCard"], div[data-name="AgentCard"]')
    n = await cards.count()
    out: List[Dict] = []

    for i in range(n):
        c = cards.nth(i)
        profile_url = ""
        name = ""
        agency = ""
        cian_exp = ""
        objects = ""

        try:
            a = c.locator('a[href^="/agents/"]').first
            href = await a.get_attribute("href") or ""
            profile_url = agents_url_from_href(href)
            name = normalize_ws(await a.inner_text())
        except Exception:
            pass

        if not profile_url:
            continue

        try:
            raw = (await c.inner_text()) or ""
            lines = [normalize_ws(x) for x in re.split(r"\n+", raw) if normalize_ws(x)]

            for ln in lines:
                lnl = ln.lower()
                if not cian_exp and "на циан" in lnl:
                    cian_exp = normalize_ws(re.sub(r"(?i)\s*на\s*циан", "", ln)).strip()
                if not objects and re.search(r"\b\d+\s+(объект|объекта|объектов)\b", lnl):
                    objects = normalize_ws(ln)

            for ln in lines:
                lnl = ln.lower()
                if "отзыв" in lnl:
                    continue
                if "на циан" in lnl:
                    continue
                if re.search(r"\b\d+\s+(объект|объекта|объектов)\b", lnl):
                    continue
                if re.fullmatch(r"\d+([.,]\d+)?", ln):
                    continue
                if ln.startswith("ID:"):
                    continue
                if len(ln) <= 80 and any(
                    w in lnl for w in ("агентство", "маклер", "риелтор", "риэлтор", "специалист", "частный", "компания")
                ):
                    agency = normalize_ws(ln)
                    break
        except Exception:
            pass

        out.append(
            {
                "profile_url": profile_url,
                "name": name,
                "agency": agency,
                "objects": objects,
                "cian": cian_exp,
            }
        )

    return out


async def collect_cards_for_region(page, region_name: str) -> Tuple[List[Dict], int]:
    ok, region_id = await select_region(page, region_name)
    if not ok or not region_id:
        return [], 0

    last_page = await get_last_page(page)
    if MAX_PAGES and MAX_PAGES > 0:
        last_page = min(last_page, MAX_PAGES)

    parsed_current = urlparse(page.url or "")
    region_domain = (
        f"{parsed_current.scheme}://{parsed_current.netloc}"
        if parsed_current.netloc else CATALOG_DOMAIN
    )

    # Если страниц много — используем параллельный сбор через несколько браузеров
    if CATALOG_BROWSERS > 1 and last_page > CATALOG_BROWSERS:
        cards = await collect_cards_parallel(region_name, region_id, region_domain, last_page, HEADLESS)
        if MAX_AGENTS and len(cards) > MAX_AGENTS:
            cards = cards[:MAX_AGENTS]
        return cards, region_id

    all_cards: List[Dict] = []
    seen: set = set()

    for page_num in range(1, last_page + 1):
        url = build_catalog_url(region_id, page_num, region_domain)
        log.info(f"{region_name} | стр.{page_num}/{last_page}: {url}")

        loaded = False
        for attempt in range(RETRY_COUNT):
            try:
                if page_num > 1:
                    clicked_pagination = False

                    # 1. Кнопка "следующая →" — реальный селектор из HTML ЦИАН
                    try:
                        next_btn = page.locator(
                            "[data-name='PaginationWrapper'] [data-testid='pagination-next']"
                        ).first
                        if await next_btn.count():
                            await next_btn.scroll_into_view_if_needed(timeout=2000)
                            await page.wait_for_timeout(random.uniform(200, 500))
                            await next_btn.click(timeout=4000)
                            await page.wait_for_load_state("domcontentloaded", timeout=12000)
                            await page.wait_for_timeout(random.uniform(200, 400))
                            clicked_pagination = True
                    except Exception:
                        pass

                    # 2. Кнопка с конкретным номером страницы
                    if not clicked_pagination:
                        try:
                            btn = page.locator(
                                "[data-name='PaginationWrapper'] [data-name='PaginationItem'] span"
                            ).filter(has_text=str(page_num)).first
                            if await btn.count():
                                await btn.scroll_into_view_if_needed(timeout=2000)
                                await page.wait_for_timeout(random.uniform(200, 500))
                                await btn.click(timeout=4000)
                                await page.wait_for_load_state("domcontentloaded", timeout=12000)
                                await page.wait_for_timeout(random.uniform(200, 400))
                                clicked_pagination = True
                        except Exception:
                            pass

                    # 3. Тихий фолбэк: прямой goto (debug — не засоряет лог)
                    if not clicked_pagination:
                        log.debug(f"{region_name} | стр.{page_num} | пагинация не найдена, goto")
                        await page.goto(url, wait_until="domcontentloaded", timeout=45_000)
                        await page.wait_for_timeout(random.uniform(300, 600))

                if not await has_catalog_cards(page, 15000):
                    raise RuntimeError("cards not found")

                await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                await page.wait_for_timeout(300)

                loaded = True
                break
            except Exception:
                log.warning(f"{region_name} | стр.{page_num} | попытка {attempt+1}/{RETRY_COUNT} неудачна")
                # При неудаче возвращаемся на стр.1 чтобы восстановить контекст
                try:
                    prev_url = build_catalog_url(region_id, max(1, page_num - 1), region_domain)
                    await page.goto(prev_url, wait_until="domcontentloaded", timeout=30_000)
                    await page.wait_for_timeout(random.uniform(2000, 4000))
                except Exception:
                    pass

        if not loaded:
            log.error(f"{region_name} | стр.{page_num} пропущена")
            continue

        cards = await extract_catalog_cards(page)
        # Если карточек нет — страница не успела загрузиться. Пробуем goto напрямую.
        if not cards:
            await page.wait_for_timeout(2000)
            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=45_000)
                await page.wait_for_timeout(1000)
            except Exception:
                pass
            cards = await extract_catalog_cards(page)

        new_cards = [c for c in cards if c.get("profile_url") and c["profile_url"] not in seen]
        for c in new_cards:
            seen.add(c["profile_url"])

        if not new_cards:
            log.warning(f"{region_name} | стр.{page_num} — карточки не найдены, пропускаю")
            continue

        all_cards.extend(new_cards)
        log.info(f"{region_name} | +{len(new_cards)} карточек (итого {len(all_cards)})")

        if MAX_AGENTS and len(all_cards) >= MAX_AGENTS:
            all_cards = all_cards[:MAX_AGENTS]
            break

        await asyncio.sleep(random.uniform(*CATALOG_DELAY))

    return all_cards, region_id


async def parse_profile_page(page, url: str) -> Dict:
    data: Dict = {
        "profile_url": url,
        "name": "",
        "phone": "",
        "email": "",
        "region_work": "",
        "work_exp": "",
        "cian": "",
        "objects": "",
        "agency": "",
    }

    try:
        await page.wait_for_selector(
            'section[data-name="AboutRealtorDesktop"], [data-name="RealtorName"]',
            timeout=15000,
        )
    except Exception:
        pass

    data["name"] = await text_or_empty(page.locator('[data-name="RealtorName"]'))

    try:
        mail = page.locator('a[href^="mailto:"]').first
        if await mail.count():
            href = await mail.get_attribute("href") or ""
            if href.startswith("mailto:"):
                data["email"] = normalize_ws(href.replace("mailto:", ""))
            if not data["email"]:
                data["email"] = normalize_ws(await mail.inner_text())
    except Exception:
        pass

    async def read_phone() -> str:
        try:
            tel = page.locator('a[href^="tel:"]').first
            if await tel.count():
                href = await tel.get_attribute("href") or ""
                if href.startswith("tel:"):
                    return clean_phone(href.replace("tel:", ""))
        except Exception:
            pass

        try:
            phone_box = page.locator('[data-name="RealtorContactsPhone"]').first
            if await phone_box.count():
                phone_text = normalize_ws(await phone_box.inner_text())
                m = re.search(r"(\+7[\d\-\sXx]{8,})", phone_text)
                if m:
                    return clean_phone(m.group(1))
        except Exception:
            pass

        return ""

    data["phone"] = await read_phone()

    if not data["phone"] or "XX" in data["phone"].upper():
        clicked = False
        selectors = [
            '[data-name="RealtorContactsPhone"] span:has-text("Показать")',
            '[data-name="RealtorContactsPhone"] button:has-text("Показать")',
            '[data-name="RealtorContactsPhone"] a:has-text("Показать")',
            '[data-name="RealtorContactsPhone"] :text("Показать")',
        ]

        for sel in selectors:
            try:
                loc = page.locator(sel).first
                if await loc.count() == 0:
                    continue
                try:
                    await loc.scroll_into_view_if_needed(timeout=3000)
                except Exception:
                    pass
                await page.wait_for_timeout(random.uniform(400, 900))
                await loc.click(timeout=5000)
                clicked = True
                break
            except Exception:
                try:
                    await loc.click(timeout=5000, force=True)
                    clicked = True
                    break
                except Exception:
                    continue

        if clicked:
            await page.wait_for_timeout(random.uniform(1500, 2500))
            for _ in range(6):
                phone_after = await read_phone()
                if phone_after:
                    data["phone"] = phone_after
                if phone_after and "XX" not in phone_after.upper():
                    break
                await page.wait_for_timeout(700)

    try:
        rows = page.locator('[data-name="DescriptionRow"]')
        rcount = await rows.count()
        for i in range(rcount):
            row = rows.nth(i)
            txt = normalize_ws(await row.inner_text())
            low = txt.lower()

            if low.startswith("регион работы") and not data["region_work"]:
                data["region_work"] = normalize_ws(re.sub(r"(?i)^регион работы", "", txt)).strip()
            elif low.startswith("агентство") and not data["agency"]:
                data["agency"] = normalize_ws(re.sub(r"(?i)^агентство", "", txt)).strip()
    except Exception:
        pass

    try:
        counters = page.locator('[data-name="Counters"]')
        if await counters.count():
            items = counters.first.locator("div")
            count = await items.count()
            for i in range(count):
                item = items.nth(i)
                txt = normalize_ws(await item.inner_text())
                if not txt:
                    continue
                low = txt.lower()

                if "опыт работы" in low and not data["work_exp"]:
                    val = normalize_ws(re.sub(r"(?i)^опыт работы", "", txt)).strip()
                    if val:
                        data["work_exp"] = val
                elif "на циан" in low and not data["cian"]:
                    val = normalize_ws(re.sub(r"(?i)^на циан", "", txt)).strip()
                    if val:
                        data["cian"] = val
                elif "в работе" in low and not data["objects"]:
                    val = normalize_ws(re.sub(r"(?i)^в работе", "", txt)).strip()
                    if val:
                        data["objects"] = val
    except Exception:
        pass

    try:
        body = await page.evaluate("document.body ? document.body.innerText : ''")
    except Exception:
        body = ""

    if body:
        if not data["email"]:
            data["email"] = extract_email_from_text(body)
        if not data["region_work"]:
            data["region_work"] = extract_region_from_text(body)
        if not data["work_exp"]:
            data["work_exp"] = extract_work_exp_from_text(body)
        if not data["cian"]:
            data["cian"] = extract_cian_from_text(body)
        if not data["objects"]:
            data["objects"] = extract_objects_from_text(body)

    if not data["name"]:
        try:
            h1 = await page.query_selector("h1")
            if h1:
                t = await safe_text(h1)
                if t and len(t) < 140:
                    data["name"] = normalize_ws(t)
        except Exception:
            pass

    for key in ("name", "agency", "objects", "email", "phone", "region_work", "work_exp", "cian"):
        data[key] = normalize_ws(data[key])

    return data


async def worker(ctx, queue: asyncio.Queue, results: list, total: int, autosave_lock: asyncio.Lock, region_name: str):
    from playwright.async_api import TimeoutError as PWTimeout

    page = await new_stealth_page(ctx)
    try:
        while True:
            try:
                idx, row = queue.get_nowait()
            except asyncio.QueueEmpty:
                break

            url = row.get("profile_url", "")
            ok = False
            for attempt in range(RETRY_COUNT):
                try:
                    await page.goto(url, wait_until="domcontentloaded", timeout=45_000)
                    ok = True
                    break
                except PWTimeout:
                    log.warning(f"[{region_name}] [{idx}/{total}] Таймаут, попытка {attempt+1}/{RETRY_COUNT}")
                    try:
                        await page.close()
                    except Exception:
                        pass
                    await asyncio.sleep(random.uniform(3, 6))
                    page = await new_stealth_page(ctx)
                except Exception as e:
                    log.warning(f"[{region_name}] [{idx}/{total}] Ошибка: {e}, попытка {attempt+1}/{RETRY_COUNT}")
                    await asyncio.sleep(3)

            if not ok:
                log.error(f"[{region_name}] [{idx}/{total}] Пропускаем {url}")
                results.append(row)
                queue.task_done()
                continue

            prof = await parse_profile_page(page, url)

            merged = dict(row)
            merged["name"] = merged.get("name") or prof.get("name") or ""
            merged["agency"] = prof.get("agency") or merged.get("agency") or ""
            merged["objects"] = prof.get("objects") or merged.get("objects") or ""
            merged["email"] = prof.get("email") or ""
            merged["phone"] = prof.get("phone") or ""
            merged["region_work"] = prof.get("region_work") or ""
            merged["work_exp"] = prof.get("work_exp") or ""
            merged["cian"] = prof.get("cian") or merged.get("cian") or ""

            results.append(merged)
            log.info(
                f"[{region_name}] [{idx}/{total}] {merged.get('name') or '—'} | "
                f"{merged.get('phone') or 'нет тел.'} | {merged.get('region_work') or '—'}"
            )

            if AUTOSAVE_EVERY and (len(results) % AUTOSAVE_EVERY == 0):
                async with autosave_lock:
                    snap = list(results)
                    await asyncio.to_thread(save_to_excel, snap, region_name, True)
                    log.info(f"[{region_name}] Автосохранение: {len(snap)} профилей")

            queue.task_done()
            await asyncio.sleep(random.uniform(0.2, 0.6))
    finally:
        await page.close()


async def enrich_profiles(cards: List[Dict], region_name: str, headless: bool) -> List[Dict]:
    from playwright.async_api import async_playwright

    results: List[Dict] = []
    total = len(cards)
    if total == 0:
        return results

    queue: asyncio.Queue = asyncio.Queue()
    for idx, row in enumerate(cards, 1):
        await queue.put((idx, row))

    autosave_lock = asyncio.Lock()

    async with async_playwright() as pw:
        browser, ctx = await make_browser_context(pw, headless)

        warmup = await new_stealth_page(ctx)
        try:
            await warmup.goto(BASE_DOMAIN_DEFAULT, wait_until="domcontentloaded", timeout=25_000)
            await warmup.wait_for_timeout(1200)
        except Exception:
            pass
        await warmup.close()

        wcount = min(PARALLEL_TABS, total)
        tasks = [
            asyncio.create_task(worker(ctx, queue, results, total, autosave_lock, region_name))
            for _ in range(wcount)
        ]
        await asyncio.gather(*tasks)
        await browser.close()

    return results


def save_to_excel(data: List[Dict], region_name: str = "", progress: bool = False) -> Path:
    tag = f"_{region_name.replace(' ', '_')}" if region_name else ""
    if progress:
        filename = OUTPUT_DIR / f"cian_realtors{tag}_PROGRESS.xlsx"
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = OUTPUT_DIR / f"cian_realtors{tag}_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Риелторы ЦИАН"

    headers = ["Ссылка", "Агентство", "Объекты", "ФИО", "Почта", "Телефон", "Регион работы", "Опыт работы", "На циане"]

    hfont = Font(bold=True, color="000000", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D0D0D0")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 20
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont
        c.alignment = center
        c.border = brd

    widths = [46, 22, 14, 24, 30, 18, 36, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    fill_none = PatternFill()
    link_font = Font(color="1A56DB", underline="single", size=10)

    def _line_count(value: str, col_width: int) -> int:
        s = str(value or "")
        if not s:
            return 1
        s = s.replace("\r\n", "\n").replace("\r", "\n")
        parts = s.split("\n")
        w = max(int(col_width), 8)
        total = 0
        for p in parts:
            p = p.strip()
            if not p:
                total += 1
                continue
            total += max(1, (len(p) + w - 1) // w)
        return max(1, total)

    for ri, row in enumerate(data, 2):
        vals = [
            row.get("profile_url", ""),
            row.get("agency", ""),
            row.get("objects", ""),
            row.get("name", ""),
            row.get("email", ""),
            row.get("phone", ""),
            row.get("region_work", ""),
            row.get("work_exp", ""),
            row.get("cian", ""),
        ]

        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill = fill_none
            c.border = brd
            c.alignment = left
            if ci == 1 and v:
                c.font = link_font
                c.hyperlink = v

        max_lines = 1
        for ci, v in enumerate(vals, 1):
            col_letter = get_column_letter(ci)
            col_width = ws.column_dimensions[col_letter].width or 10
            max_lines = max(max_lines, _line_count(v, int(col_width)))
        ws.row_dimensions[ri].height = min(240, 14 * max_lines + 4)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(data)+1}"
    wb.save(filename)
    log.info(f"Файл сохранён: {filename}")
    return filename


async def choose_targets_via_ui(headless: bool) -> List[str]:
    from playwright.async_api import async_playwright

    async with async_playwright() as pw:
        browser, ctx = await make_browser_context(pw, headless)
        page = await new_stealth_page(ctx)

        ok = await ensure_catalog_open(page)
        if not ok:
            await browser.close()
            return []

        names = await get_all_region_names(page)
        await browser.close()
        return names


def choose_mode(region_names: List[str], cli_from: int = 0, cli_to: int = 0) -> Tuple[str, List[str]]:
    # --- Режим CLI: --from N [--to M] ---
    if cli_from > 0:
        start_idx = max(1, cli_from)
        end_idx   = cli_to if cli_to > 0 else len(region_names)
        end_idx   = min(end_idx, len(region_names))
        selected  = region_names[start_idx - 1 : end_idx]
        mode_name = f"С {start_idx} по {end_idx} регион"
        return mode_name, selected

    print("\n" + "=" * 58)
    print("  Выберите режим:")
    print("=" * 58)
    print("  [0]  Все регионы")
    print("  [1]  Один регион")
    print("  [2]  Диапазон регионов (с N-го по M-й)")
    print("=" * 58)

    while True:
        choice = input("  Введите номер: ").strip()
        if choice == "0":
            return "Все регионы", region_names
        if choice == "1":
            print("\n  Доступные регионы:")
            for i, name in enumerate(region_names, 1):
                print(f"  [{i:>3}] {name}")
            while True:
                raw = input("\n  Введите номер региона: ").strip()
                if raw.isdigit():
                    idx = int(raw)
                    if 1 <= idx <= len(region_names):
                        return region_names[idx - 1], [region_names[idx - 1]]
                print("  Неверный выбор.")
        if choice == "2":
            print("\n  Доступные регионы:")
            for i, name in enumerate(region_names, 1):
                print(f"  [{i:>3}] {name}")
            while True:
                raw_from = input(f"\n  С какого номера (1–{len(region_names)}): ").strip()
                raw_to   = input(f"  По какой номер   (1–{len(region_names)}, Enter = до конца): ").strip()
                if raw_from.isdigit():
                    f = int(raw_from)
                    t = int(raw_to) if raw_to.isdigit() else len(region_names)
                    if 1 <= f <= t <= len(region_names):
                        selected = region_names[f - 1 : t]
                        return f"С {f} по {t} регион", selected
                print("  Неверный диапазон.")
        print("  Неверный выбор.")



async def collect_cards_worker(pw, headless: bool, worker_id: int, n_workers: int,
                               region_name: str, region_id: int, region_domain: str,
                               last_page: int, start_delay: float) -> List[Dict]:
    """
    Один воркер: листает каталог с самого начала кликами,
    но собирает только каждую n_workers-ю страницу начиная с worker_id.
    Например, 3 воркера: первый берёт стр.1,4,7,10..., второй 2,5,8,11..., третий 3,6,9,12...
    Так все три идут параллельно без прыжков в середину каталога.
    """
    all_cards: List[Dict] = []
    seen: set = set()

    # Разнесённый старт чтобы не все 3 браузера открывались одновременно
    if start_delay > 0:
        await asyncio.sleep(start_delay)

    browser, ctx = await make_browser_context(pw, headless)
    page = await new_stealth_page(ctx)

    warmup_url = build_catalog_url(region_id, 1, region_domain)
    try:
        await page.goto(warmup_url, wait_until="domcontentloaded", timeout=35_000)
        await page.wait_for_timeout(random.uniform(1200, 2000))
        if not await has_catalog_cards(page, 10000):
            log.warning(f"{region_name} [w{worker_id}] | прогрев не удался")
            await browser.close()
            return []
    except Exception as e:
        log.warning(f"{region_name} [w{worker_id}] | ошибка прогрева: {e}")
        await browser.close()
        return []

    current_page = 1
    my_pages = list(range(worker_id, last_page + 1, n_workers))
    log.info(f"{region_name} [w{worker_id}] | буду собирать {len(my_pages)} стр. из {last_page}")

    for target_page in my_pages:
        # Листаем кликами до нужной страницы
        while current_page < target_page:
            try:
                nb = page.locator("[data-name='PaginationWrapper'] [data-testid='pagination-next']").first
                if await nb.count():
                    await nb.scroll_into_view_if_needed(timeout=2000)
                    await page.wait_for_timeout(random.uniform(150, 350))
                    await nb.click(timeout=4000)
                    await page.wait_for_load_state("domcontentloaded", timeout=12000)
                    await page.wait_for_timeout(random.uniform(300, 600))
                    current_page += 1
                else:
                    log.debug(f"{region_name} [w{worker_id}] | кнопка след. не найдена на стр.{current_page}")
                    break
            except Exception:
                break

        if current_page != target_page:
            log.warning(f"{region_name} [w{worker_id}] | не удалось долистать до стр.{target_page}, пропускаю")
            continue

        if not await has_catalog_cards(page, 8000):
            log.warning(f"{region_name} [w{worker_id}] | нет карточек на стр.{target_page}")
            continue

        for _ in range(2):
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(400)

        cards = await extract_catalog_cards(page)
        new_cards = [c for c in cards if c.get("profile_url") and c["profile_url"] not in seen]
        for c in new_cards:
            seen.add(c["profile_url"])
        all_cards.extend(new_cards)

        if new_cards:
            log.info(f"{region_name} [w{worker_id}] стр.{target_page} +{len(new_cards)} карточек")

        await asyncio.sleep(random.uniform(*CATALOG_DELAY))

    await browser.close()
    log.info(f"{region_name} [w{worker_id}] | завершён, собрано {len(all_cards)} карточек")
    return all_cards


async def collect_cards_parallel(region_name: str, region_id: int, region_domain: str,
                                  last_page: int, headless: bool) -> List[Dict]:
    """
    Запускает CATALOG_BROWSERS воркеров.
    Прогрев — строго последовательный (один за другим с паузой),
    чтобы ЦИАН не заблокировал несколько новых сессий сразу.
    После успешного прогрева все воркеры идут параллельно.
    """
    from playwright.async_api import async_playwright

    n = min(CATALOG_BROWSERS, last_page)
    log.info(f"{region_name} | параллельный сбор: {n} браузеров, всего стр.: {last_page}")

    # asyncio.Event для каждого воркера — сигнал "я прогрелся, следующий может стартовать"
    warmup_events = [asyncio.Event() for _ in range(n)]
    # Первый воркер стартует сразу
    warmup_events[0].set()

    async def collect_cards_worker_gated(pw, i):
        # Ждём пока предыдущий воркер прогреется
        await warmup_events[i].wait()
        browser, ctx = await make_browser_context(pw, headless)
        page = await new_stealth_page(ctx)
        # Прогрев: открываем базовый каталог и выбираем регион через модалку,
        # точно так же как основной браузер — иначе ЦИАН отдаёт главную страницу
        try:
            ok = await ensure_catalog_open(page)
            if not ok:
                raise RuntimeError("catalog not opened")
            ok2, _ = await select_region(page, region_name)
            if not ok2:
                raise RuntimeError("region not selected")
            if not await has_catalog_cards(page, 10000):
                raise RuntimeError("no cards after region select")
        except Exception as e:
            log.warning(f"{region_name} [w{i+1}] | прогрев не удался: {e}")
            await browser.close()
            if i + 1 < n:
                await asyncio.sleep(random.uniform(5, 8))
                warmup_events[i + 1].set()
            return []

        log.info(f"{region_name} [w{i+1}] | прогрев OK, стартую сбор")
        log.info(f"{region_name} [w{i+1}] | URL после прогрева: {page.url}")

        # Разблокируем следующий воркер с паузой
        if i + 1 < n:
            await asyncio.sleep(random.uniform(8, 12))
            warmup_events[i + 1].set()

        # Основной сбор — каждый берёт свои страницы через шаг n
        all_cards: List[Dict] = []
        seen: set = set()
        current_page = 1
        my_pages = list(range(i + 1, last_page + 1, n))
        log.info(f"{region_name} [w{i+1}] | собираю {len(my_pages)} стр.")

        for target_page in my_pages:
            # Листаем до нужной страницы кликами
            while current_page < target_page:
                try:
                    # Скролл вниз чтобы пагинация появилась (lazy render)
                    await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                    await page.wait_for_timeout(600)

                    nb = page.locator("[data-name='PaginationWrapper'] [data-testid='pagination-next']").first
                    await nb.wait_for(state="visible", timeout=5000)
                    await nb.scroll_into_view_if_needed(timeout=2000)
                    await page.wait_for_timeout(random.uniform(150, 350))
                    await nb.click(timeout=4000)
                    await page.wait_for_load_state("domcontentloaded", timeout=15000)
                    await page.wait_for_timeout(random.uniform(400, 700))
                    current_page += 1
                except Exception:
                    # Логируем для диагностики и прерываем
                    log.debug(f"{region_name} [w{i+1}] | кнопка не найдена на стр.{current_page}, цель={target_page}")
                    break

            if current_page != target_page:
                log.warning(f"{region_name} [w{i+1}] | не долистал до стр.{target_page} (на стр.{current_page}), пропускаю")
                continue

            # Скролл для загрузки карточек
            for _ in range(3):
                await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                await page.wait_for_timeout(500)

            if not await has_catalog_cards(page, 8000):
                log.warning(f"{region_name} [w{i+1}] | нет карточек на стр.{target_page}")
                continue

            cards = await extract_catalog_cards(page)
            new_cards = [c for c in cards if c.get("profile_url") and c["profile_url"] not in seen]
            for c in new_cards:
                seen.add(c["profile_url"])
            all_cards.extend(new_cards)
            if new_cards:
                log.info(f"{region_name} [w{i+1}] стр.{target_page} +{len(new_cards)} карточек")

            # Синхронизируем current_page из URL
            real_page = parse_qs(urlparse(page.url).query).get("page", [None])[0]
            if real_page and real_page.isdigit():
                current_page = int(real_page)

            await asyncio.sleep(random.uniform(*CATALOG_DELAY))

        await browser.close()
        log.info(f"{region_name} [w{i+1}] | завершён, собрано {len(all_cards)} карточек")
        return all_cards

    async with async_playwright() as pw:
        results = await asyncio.gather(*[collect_cards_worker_gated(pw, i) for i in range(n)])

    seen: set = set()
    all_cards: List[Dict] = []
    for chunk in results:
        for c in chunk:
            u = c.get("profile_url", "")
            if u and u not in seen:
                seen.add(u)
                all_cards.append(c)

    log.info(f"{region_name} | итого уникальных карточек: {len(all_cards)}")
    return all_cards


async def collect_all_regions_cards(selected_regions: List[str], headless: bool) -> Dict[str, List[Dict]]:
    """Собирает каталог одного региона за раз — не держит все регионы в памяти."""
    from playwright.async_api import async_playwright

    result: Dict[str, List[Dict]] = {}

    for i, region_name in enumerate(selected_regions, 1):
        log.info(f"=== Каталог {i}/{len(selected_regions)}: {region_name} ===")
        async with async_playwright() as pw:
            browser, ctx = await make_browser_context(pw, headless)
            page = await new_stealth_page(ctx)

            ok = await ensure_catalog_open(page)
            if not ok:
                log.warning(f"{region_name} | не удалось открыть каталог, пропускаю")
                await browser.close()
                result[region_name] = []
                continue

            cards, region_id = await collect_cards_for_region(page, region_name)
            await browser.close()

        if cards:
            log.info(f"{region_name} | regionId={region_id} | собрано карточек: {len(cards)}")
        else:
            log.warning(f"{region_name} | regionId={region_id} | карточки не собраны")
        result[region_name] = cards

        if i < len(selected_regions):
            pause = random.uniform(*REGION_PAUSE)
            log.info(f"Пауза перед следующим регионом: {pause:.1f} сек.")
            await asyncio.sleep(pause)

    return result


def main():
    parser = argparse.ArgumentParser(description="Парсер риелторов ЦИАН")
    parser.add_argument(
        "--from", dest="from_region", type=int, default=0,
        metavar="N",
        help="Номер региона, с которого начинать (1-based). "
             "При задании этого флага интерактивные паузы отключаются.",
    )
    parser.add_argument(
        "--to", dest="to_region", type=int, default=0,
        metavar="M",
        help="Номер региона, которым заканчивать (включительно). "
             "По умолчанию — последний регион.",
    )
    args = parser.parse_args()
    auto_mode = args.from_region > 0  # True => без input()-пауз

    print("=" * 60)
    print("  Парсер риелторов ЦИАН  |  v11.0")
    print("=" * 60)

    if not STEALTH_AVAILABLE:
        print("\n  ⚠️  playwright-stealth не установлен")
        print("  Установите: pip install playwright-stealth")
        if not auto_mode:
            input("\n  Нажмите Enter для продолжения...\n")
    else:
        print("\n  ✅ playwright-stealth активен\n")

    print("  Получаю список регионов из UI каталога...")
    region_names = asyncio.run(choose_targets_via_ui(HEADLESS))

    if not region_names:
        print("\n  ⚠️  Не удалось получить список регионов.")
        if not auto_mode:
            input("\nНажмите Enter для выхода...")
        return

    mode_name, selected_regions = choose_mode(region_names, args.from_region, args.to_region)

    print(f"\n  Режим         : {mode_name}")
    print(f"  Регионов      : {len(selected_regions)}")
    print(f"  Параллельность: {PARALLEL_TABS} вкладок")
    print(f"  Страниц       : {'все' if not MAX_PAGES else MAX_PAGES}")
    print(f"  Автосейв      : каждые {AUTOSAVE_EVERY} профилей")
    print("=" * 60)
    if not auto_mode:
        input("\n  Нажмите Enter для старта...\n")
    else:
        print("\n  Авторежим: старт без ожидания ввода\n")

    start = time.time()
    all_results: List[Dict] = []

    async def run_one_region(region_name: str, idx_region: int) -> List[Dict]:
        """Каталог + профили одного региона за один проход."""
        print(f"\n▶ Регион {idx_region}/{len(selected_regions)}: {region_name}")

        from playwright.async_api import async_playwright
        async with async_playwright() as pw:
            browser, ctx = await make_browser_context(pw, HEADLESS)
            page = await new_stealth_page(ctx)
            ok = await ensure_catalog_open(page)
            if not ok:
                print(f"  ⚠️  Не удалось открыть каталог для {region_name}")
                await browser.close()
                return []
            cards, region_id = await collect_cards_for_region(page, region_name)
            await browser.close()

        if not cards:
            print(f"  ⚠️  Карточки не найдены. Проверьте cian_parser.log")
            return []

        if MAX_AGENTS and len(cards) > MAX_AGENTS:
            cards = cards[:MAX_AGENTS]

        tag = sanitize_tag(region_name)
        lf = OUTPUT_DIR / f"profile_links_{tag}.txt"
        lf.write_text("\n".join([c["profile_url"] for c in cards if c.get("profile_url")]), encoding="utf-8")
        log.info(f"Ссылки сохранены: {lf.name} ({len(cards)} шт.)")

        print(f"  Карточек каталога: {len(cards)}")
        print(f"  Шаг 2/2: парсинг {len(cards)} профилей ({PARALLEL_TABS} вкладок)...")
        results = await enrich_profiles(cards, region_name, HEADLESS)

        if results:
            out = save_to_excel(results, region_name, progress=False)
            print(f"  ✅ {region_name}: {len(results)} строк → {out.name}")
        else:
            print(f"  ⚠️  {region_name}: данные профилей не собраны.")

        return results

    async def run_all():
        for idx_region, region_name in enumerate(selected_regions, 1):
            try:
                results = await run_one_region(region_name, idx_region)
                all_results.extend(results)
            except Exception as e:
                log.error(f"[{region_name}] критическая ошибка: {e}")

            if idx_region < len(selected_regions):
                pause = random.uniform(*REGION_PAUSE)
                print(f"  Пауза перед следующим регионом: {pause:.1f} сек.")
                await asyncio.sleep(pause)

    try:
        asyncio.run(run_all())

    except KeyboardInterrupt:
        print("\n⛔ Остановлено (Ctrl+C). Сохраняю собранное...")

    finally:
        if len(selected_regions) > 1 and all_results:
            out_all = save_to_excel(all_results, "ALL", progress=False)
            print(f"\n✅ Общий файл: {out_all}")

        elapsed = time.time() - start
        if all_results:
            print(f"\nИтого: {len(all_results)} строк за {elapsed:.0f} сек.")
            print(f"(~{elapsed/max(len(all_results),1):.1f} сек/строка)")
        else:
            print("\n⚠️  Данные не собраны. Проверьте cian_parser.log")

    if not auto_mode:
        input("\nНажмите Enter для выхода...")


if __name__ == "__main__":
    main()