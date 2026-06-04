import argparse
import asyncio
import logging
import math
import random
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional
from urllib.parse import parse_qs, urljoin, urlparse

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from playwright_stealth import Stealth as _Stealth

    async def _apply_stealth(page):
        await _Stealth().apply_stealth_async(page)

    STEALTH_AVAILABLE = True
except ImportError:
    try:
        from playwright_stealth import stealth_async as _apply_stealth  # type: ignore

        STEALTH_AVAILABLE = True
    except ImportError:

        async def _apply_stealth(page):
            return None

        STEALTH_AVAILABLE = False


OUTPUT_DIR = Path(__file__).parent
BASE_DOMAIN_DEFAULT = "https://www.cian.ru"
RETRY_COUNT = 3
CATALOG_DELAY = (0.4, 0.8)
PROFILE_DELAY = (0.5, 1.0)
REGION_PAUSE = (10.0, 20.0)

REGIONS = {
    1: "Москва",
    2: "Санкт-Петербург",
    4630: "Челябинская область",
}

FALLBACK_REGION_NAMES = [
    "Москва",
    "Санкт-Петербург",
    "Адыгея",
    "Алтай (Республика)",
    "Алтайский край",
    "Амурская область",
    "Архангельская область",
    "Астраханская область",
    "Башкортостан",
    "Белгородская область",
    "Брянская область",
    "Бурятия",
    "Владимирская область",
    "Волгоградская область",
    "Вологодская область",
    "Воронежская область",
    "Дагестан",
    "Еврейская АО",
    "Ивановская область",
    "Ингушетия",
    "Иркутская область",
    "Кабардино-Балкария",
    "Калининградская область",
    "Калмыкия",
    "Калужская область",
    "Камчатский край",
    "Карачаево-Черкесия",
    "Карелия",
    "Кемеровская область",
    "Кировская область",
    "Коми",
    "Костромская область",
    "Краснодарский край",
    "Красноярский край",
    "Курганская область",
    "Курская область",
    "Ленинградская область",
    "Липецкая область",
    "Магаданская область",
    "Марий Эл",
    "Мордовия",
    "Московская область",
    "Мурманская область",
    "Ненецкий АО",
    "Нижегородская область",
    "Новгородская область",
    "Новосибирская область",
    "Омская область",
    "Оренбургская область",
    "Орловская область",
    "Пензенская область",
    "Пермский край",
    "Приморский край",
    "Псковская область",
    "Ростовская область",
    "Рязанская область",
    "Самарская область",
    "Саратовская область",
    "Саха (Якутия)",
    "Сахалинская область",
    "Свердловская область",
    "Северная Осетия-Алания",
    "Смоленская область",
    "Ставропольский край",
    "Тамбовская область",
    "Татарстан",
    "Тверская область",
    "Томская область",
    "Тульская область",
    "Тыва",
    "Тюменская область",
    "Удмуртия",
    "Ульяновская область",
    "Хабаровский край",
    "Хакасия",
    "Ханты-Мансийский АО",
    "Челябинская область",
    "Чечня",
    "Чувашия",
    "Чукотский АО",
    "Ямало-Ненецкий АО",
    "Ярославская область",
    "Крым",
    "Севастополь",
    "Забайкальский край",
]

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(OUTPUT_DIR / "cian_agencies_parser.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

DEBUG_LOG_FILE = OUTPUT_DIR / f"cian_agencies_debug_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
debug_log = logging.getLogger("cian_agencies_debug")
debug_log.setLevel(logging.DEBUG)
debug_log.propagate = False
_debug_handler = logging.FileHandler(DEBUG_LOG_FILE, encoding="utf-8")
_debug_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
debug_log.addHandler(_debug_handler)

SENSITIVE_DEBUG_FIELDS = {
    "cookie",
    "cookies",
    "headers",
    "token",
    "tokens",
    "localStorage",
    "sessionStorage",
    "html",
    "content",
}


def debug_step(message: str, **fields) -> None:
    safe_fields = {}
    for key, value in fields.items():
        if key in SENSITIVE_DEBUG_FIELDS:
            continue
        text = re.sub(r"\s+", " ", str(value or "")).strip()
        if len(text) > 500:
            text = text[:500] + "..."
        safe_fields[key] = text

    if safe_fields:
        details = " | ".join(f"{key}={value}" for key, value in safe_fields.items())
        debug_log.info("%s | %s", message, details)
    else:
        debug_log.info("%s", message)


async def new_stealth_page(ctx):
    page = await ctx.new_page()
    await _apply_stealth(page)
    return page


def normalize_ws(value: str) -> str:
    value = (value or "").replace("\xa0", " ").replace("\u2009", " ")
    return re.sub(r"\s+", " ", value).strip()


def sanitize_tag(value: str) -> str:
    value = re.sub(r"[^\w\-]+", "_", (value or "").strip(), flags=re.U)
    value = re.sub(r"_+", "_", value)
    return value.strip("_") or "region"


def build_catalog_url(region_id: int, page_num: int = 1, base_domain: str = "https://cian.ru") -> str:
    return f"{base_domain.rstrip('/')}/agentstva/?regionId={region_id}&page={page_num}"


def parse_region_id_from_url(url: str) -> int:
    try:
        query = parse_qs(urlparse(url).query)
        raw = query.get("regionId", ["0"])[0]
        return int(raw)
    except Exception:
        return 0


def agency_url_from_href(href: str) -> str:
    if not href:
        return ""

    href = href.strip()
    if href.startswith("//"):
        href = "https:" + href
    href = urljoin(BASE_DOMAIN_DEFAULT, href)
    parsed = urlparse(href)
    path = re.sub(r"/+", "/", parsed.path or "")

    match = re.search(r"/company/(\d+)", path)
    if not match:
        return ""

    return f"{BASE_DOMAIN_DEFAULT}/company/{match.group(1)}/"


def agency_url_from_id(agency_id) -> str:
    agency_id = str(agency_id or "").strip()
    if not agency_id or not agency_id.isdigit():
        return ""
    return f"{BASE_DOMAIN_DEFAULT}/company/{agency_id}/"


def clean_phone(raw: str) -> str:
    raw = normalize_ws(raw)
    if not raw:
        return ""

    if "X" in raw.upper():
        match = re.search(r"(\+?7[\d\-\sXx]{8,})", raw)
        if match:
            value = normalize_ws(match.group(1))
            return value if value.startswith("+") else "+" + value
        return raw

    digits = re.sub(r"\D", "", raw)
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    if len(digits) == 10:
        digits = "7" + digits
    if len(digits) >= 11:
        return f"+{digits[0]}-{digits[1:4]}-{digits[4:7]}-{digits[7:9]}-{digits[9:11]}"
    return raw


def append_unique(values: List[str], value: str) -> None:
    value = normalize_ws(value)
    if value and value not in values:
        values.append(value)


def join_unique(values: List[str]) -> str:
    out: List[str] = []
    for value in values:
        append_unique(out, value)
    return "\n".join(out)


def split_multiline_values(value: str) -> List[str]:
    return [normalize_ws(part) for part in (value or "").splitlines() if normalize_ws(part)]


def normalize_multiline(value: str) -> str:
    return "\n".join(split_multiline_values(value))


def extract_objects_from_text(text: str) -> str:
    match = re.search(r"(\d[\d\s]*)\s*(объект|объекта|объектов|предложение|предложения|предложений)\b", text or "", re.I)
    return normalize_ws(match.group(0)) if match else ""


def normalize_offers_count(offers_count) -> str:
    if offers_count is None:
        return ""

    value = normalize_ws(str(offers_count))
    if not value:
        return ""

    if re.fullmatch(r"\d[\d\s]*", value):
        return f"{value} предложений"

    return value


def extract_reviews_from_text(text: str) -> str:
    match = re.search(r"(\d[\d\s]*)\s*(отзыв|отзыва|отзывов)\b", text or "", re.I)
    return normalize_ws(match.group(0)) if match else ""


def extract_agents_count_from_text(text: str) -> str:
    match = re.search(r"(\d[\d\s]*)\s*(сотрудник|сотрудника|сотрудников|агент|агента|агентов)\b", text or "", re.I)
    return normalize_ws(match.group(0)) if match else ""


def normalize_cian_term(value: str) -> str:
    value = normalize_ws(str(value or ""))
    if not value:
        return ""
    return normalize_ws(re.sub(r"\s+на\s+Циан\s*$", "", value, flags=re.I))


def extract_cian_term_from_text(text: str) -> str:
    match = re.search(
        r"((?:\d+\s*(?:год|года|лет)(?:\s+\d+\s*(?:месяц|месяца|месяцев))?)|(?:\d+\s*(?:месяц|месяца|месяцев)))\s+на\s+Циан\b",
        text or "",
        re.I,
    )
    return normalize_cian_term(match.group(1)) if match else ""


def extract_labeled_value(text: str, label: str) -> str:
    lines = [normalize_ws(line) for line in (text or "").splitlines()]
    lines = [line for line in lines if line]
    label_norm = label.rstrip(":").lower()

    for idx, line in enumerate(lines):
        line_clean = line.rstrip(":")
        if line_clean.lower() == label_norm:
            if idx + 1 < len(lines):
                return normalize_labeled_value(label_norm, lines[idx + 1])
            return ""
        if line.lower().startswith(label_norm + ":"):
            value = normalize_ws(line.split(":", 1)[1])
            if value:
                return normalize_labeled_value(label_norm, value)
            if idx + 1 < len(lines):
                return normalize_labeled_value(label_norm, lines[idx + 1])
    return ""


def extract_labeled_values(text: str, label: str) -> List[str]:
    values: List[str] = []
    lines = [normalize_ws(line) for line in (text or "").splitlines()]
    lines = [line for line in lines if line]
    label_norm = label.rstrip(":").lower()

    for idx, line in enumerate(lines):
        line_clean = line.rstrip(":")
        value = ""

        if line_clean.lower() == label_norm and idx + 1 < len(lines):
            value = lines[idx + 1]
        elif line.lower().startswith(label_norm + ":"):
            value = normalize_ws(line.split(":", 1)[1])
            if not value and idx + 1 < len(lines):
                value = lines[idx + 1]

        value = normalize_labeled_value(label_norm, value)
        if value:
            append_unique(values, value)

    return values


def normalize_labeled_value(label: str, value: str) -> str:
    value = normalize_ws(value)
    if not value:
        return ""
    low = value.lower()
    if label == "сайт":
        if value.startswith("//"):
            value = value[2:]
        if value.startswith("http://"):
            value = value.replace("http://", "", 1)
        if value.startswith("https://"):
            value = value.replace("https://", "", 1)

        if not is_valid_agency_site(value):
            return ""

    return value


def is_valid_agency_site(value: str) -> bool:
    value = normalize_ws(value)
    if not value:
        return False
    if "@" in value or re.search(r"\s", value):
        return False

    candidate = value if re.match(r"^[a-z][a-z0-9+.-]*://", value, re.I) else f"//{value}"
    parsed = urlparse(candidate)
    host = (parsed.hostname or "").lower().strip(".")
    if host.startswith("www."):
        host = host[4:]

    blocked_hosts = {
        "cian.ru",
        "ciangroup.ru",
        "ir.ciangroup.ru",
        "onelink.me",
        "cianag.onelink.me",
        "cian-gp.onelink.me",
        "cian-appstore.onelink.me",
        "cian-rustore.onelink.me",
    }
    if not host or any(host == blocked or host.endswith(f".{blocked}") for blocked in blocked_hosts):
        return False

    return bool(re.fullmatch(r"[a-zа-яё0-9-]+(?:\.[a-zа-яё0-9-]+)+", host, re.I))


def normalize_site_value(value: str) -> str:
    value = normalize_ws(value).strip(" \t\r\n.,;)")
    if value.startswith("//"):
        value = value[2:]
    return value


def site_identity(value: str) -> str:
    value = normalize_site_value(value).lower()
    value = re.sub(r"^https?://", "", value)
    value = re.sub(r"^www\.", "", value)
    return value.rstrip("/")


def append_unique_site(values: List[str], value: str) -> None:
    value = normalize_site_value(value)
    if not value:
        return
    identity = site_identity(value)
    if identity and all(site_identity(existing) != identity for existing in values):
        values.append(value)


def extract_phones_from_text(text: str) -> str:
    matches = re.findall(r"(\+?\s*[78][\s().\-]*(?:[\dXx][\s().\-]*){10})", text or "")
    phones: List[str] = []

    for match in matches:
        phone = clean_phone(match)
        if "XX" in phone.upper():
            continue
        if phone and phone not in phones:
            phones.append(phone)

    return "\n".join(phones)


def profile_has_enrichment(data: Dict) -> bool:
    return any(data.get(key) for key in ("phone", "site", "region", "objects", "rating", "cian"))


def normalize_status(value: str) -> str:
    value = normalize_ws(value)
    if value == "ok":
        return "Успешно"
    if value == "partial":
        return "Частично"
    return value


async def make_browser_context(pw, headless: bool):
    browser = await pw.chromium.launch(
        headless=headless,
        args=[
            "--no-sandbox",
            "--disable-blink-features=AutomationControlled",
            "--disable-infobars",
            "--disable-dev-shm-usage",
            "--disable-extensions",
        ],
    )
    ctx = await browser.new_context(
        viewport={"width": 1440, "height": 1000},
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        locale="ru-RU",
        timezone_id="Europe/Moscow",
        extra_http_headers={
            "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
            "sec-ch-ua": '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
        },
    )
    await ctx.add_init_script(
        "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
        "window.chrome={runtime:{}};"
        "Object.defineProperty(navigator,'plugins',{get:()=>[1,2,3,4,5]});"
        "Object.defineProperty(navigator,'languages',{get:()=>['ru-RU','ru','en-US','en']});"
    )

    blocked_domains = (
        "google-analytics.com",
        "googletagmanager.com",
        "mc.yandex.ru",
        "top-fwz1.mail.ru",
        "counter.yadro.ru",
        "hotjar.com",
        "doubleclick.net",
        "facebook.net",
        "vk.com/rtrg",
        "sentry.io",
        "datadome.co",
        "adfox.ru",
    )

    async def route_handler(route):
        try:
            resource_type = route.request.resource_type
            request_url = route.request.url
            if resource_type in ("image", "media", "font", "stylesheet"):
                await route.abort()
                return
            if any(domain in request_url for domain in blocked_domains):
                await route.abort()
                return
        except Exception:
            pass
        await route.continue_()

    await ctx.route("**/*", route_handler)
    return browser, ctx


async def text_or_empty(locator) -> str:
    try:
        if await locator.count() == 0:
            return ""
        return normalize_ws(await locator.first.inner_text())
    except Exception:
        return ""


async def get_agents_listing_state(page) -> Dict:
    try:
        state = await page.evaluate(
            """
            () => {
              const cfg = window._cianConfig && window._cianConfig['agents-directory-frontend'];
              if (!Array.isArray(cfg)) return {};
              const item = cfg.find(x => x && x.key === 'initialState');
              return item && item.value ? item.value : {};
            }
            """
        )
        return state or {}
    except Exception:
        return {}


async def has_agency_cards(page, timeout: int = 7000) -> bool:
    agency_card_selectors = [
        '[data-name="AgencyCard"]',
        'div[data-name="AgencyCard"]',
    ]

    for selector in agency_card_selectors:
        try:
            await page.wait_for_selector(selector, timeout=timeout)
            if await page.locator(selector).count() > 0:
                return True
        except Exception:
            pass

    try:
        found = await page.evaluate(
            """
            () => Array.from(document.querySelectorAll('a[href]')).some(a => {
              const href = a.getAttribute('href') || '';
              return /\\/company\\/\\d+/.test(href) && !/\\/agents\\//.test(href);
            })
            """
        )
        return bool(found)
    except Exception:
        return False


async def click_any(page, selectors: List[str], timeout: int = 4500) -> bool:
    for selector in selectors:
        try:
            locators = page.locator(selector)
            count = await locators.count()
            for i in range(min(count, 6)):
                locator = locators.nth(i)
                try:
                    await locator.scroll_into_view_if_needed(timeout=timeout)
                except Exception:
                    pass
                try:
                    await page.wait_for_timeout(random.uniform(300, 700))
                    await locator.click(timeout=timeout)
                    await page.wait_for_timeout(random.uniform(1200, 2200))
                    return True
                except Exception:
                    try:
                        await locator.click(timeout=timeout, force=True)
                        await page.wait_for_timeout(random.uniform(1200, 2200))
                        return True
                    except Exception:
                        continue
        except Exception:
            continue
    return False


async def ensure_agencies_catalog_open(page) -> bool:
    warm_url = build_catalog_url(4630, 1)
    for _ in range(RETRY_COUNT):
        try:
            log.info("Открываю каталог агентств для выбора региона: %s", warm_url)
            await page.goto(warm_url, wait_until="domcontentloaded", timeout=45_000)
            await page.wait_for_timeout(random.uniform(1400, 2400))
            if await has_agency_cards(page, 12_000):
                return True

            body = ""
            try:
                body = await page.evaluate("document.body ? document.body.innerText : ''")
            except Exception:
                body = ""

            if "агентств" in (body or "").lower():
                clicked = await click_any(
                    page,
                    [
                        "a[href*='/agentstva/']",
                        "button:has-text('Выбрать агентство')",
                        "a:has-text('Выбрать агентство')",
                        "button:has-text('Найти агентство')",
                        "a:has-text('Найти агентство')",
                    ],
                )
                if clicked and await has_agency_cards(page, 15_000):
                    return True
        except Exception:
            await page.wait_for_timeout(random.uniform(2000, 4000))
    return False


async def open_region_modal(page) -> bool:
    selectors = [
        "button[data-name='GeoLocationButton']",
        "[data-name='GeoLocationButton']",
        "button:has-text('Москва')",
        "button:has-text('Челябинская область')",
        "button:has-text('Санкт-Петербург')",
        "header button",
    ]

    for selector in selectors:
        try:
            locators = page.locator(selector)
            count = await locators.count()

            for i in range(min(count, 5)):
                button = locators.nth(i)

                try:
                    text = normalize_ws(await button.inner_text(timeout=1000))
                except Exception:
                    text = ""

                if selector == "header button" and not any(
                    word in text.lower()
                    for word in ["москва", "санкт", "челябин", "регион"]
                ):
                    continue

                try:
                    await button.scroll_into_view_if_needed(timeout=3000)
                except Exception:
                    pass

                try:
                    await button.click(timeout=5000)
                except Exception:
                    await button.click(timeout=5000, force=True)

                await page.wait_for_timeout(1000)

                if await page.locator("text=Выберите регион").count() > 0:
                    return True

                if await page.locator("div[data-name='PopularRegionsDesktop'], div[data-name='SpecialRegionsDesktop']").count() > 0:
                    return True

        except Exception:
            continue

    return False

def is_region_name(value: str) -> bool:
    value = normalize_ws(value)
    if not value:
        return False

    bad_words = [
        "выберите регион",
        "сохранить",
        "найти",
        "войти",
        "разместить",
        "аренда",
        "продажа",
        "новостройки",
        "ипотека",
        "сервисы",
    ]

    low = value.lower()
    if any(word in low for word in bad_words):
        return False

    allowed_exact = {
        "Москва",
        "Санкт-Петербург",
        "Адыгея",
        "Алтай (Республика)",
        "Алтайский край",
        "Крым",
        "Севастополь",
        "Забайкальский край",
    }

    if value in allowed_exact:
        return True

    region_words = [
        "область",
        "край",
        "республика",
        "ао",
        "округ",
        "башкортостан",
        "бурятия",
        "дагестан",
        "ингушетия",
        "калмыкия",
        "карелия",
        "коми",
        "марий эл",
        "мордовия",
        "саха",
        "татарстан",
        "тыва",
        "удмуртия",
        "хакасия",
        "чечня",
        "чувашия",
    ]

    return any(word in low for word in region_words)

async def get_all_region_names(page) -> List[str]:
    ok = await open_region_modal(page)
    if not ok:
        log.warning("Не удалось открыть модальное окно выбора региона")
        await debug_agencies_page(page)
        return []

    names: List[str] = []

    selectors = [
        "div[data-name='SpecialRegionsDesktop'] label span",
        "div[data-name='PopularRegionsDesktop'] p",
        "div[data-name='SpecialRegionsDesktop'] span",
        "div[data-name='PopularRegionsDesktop'] span",
        "label span",
        "button span",
        "p",
    ]

    for selector in selectors:
        try:
            locators = page.locator(selector)
            count = await locators.count()

            for i in range(count):
                name = normalize_ws(await locators.nth(i).inner_text())

                if not is_region_name(name):
                    continue

                if name not in names:
                    names.append(name)
        except Exception:
            continue

    if not names:
        log.warning("Модальное окно открылось, но регионы не найдены")
        await debug_agencies_page(page)

    return names


async def get_selected_region_name(page) -> str:
    return await text_or_empty(page.locator("button[data-name='GeoLocationButton'] span"))


async def select_region(page, region_name: str) -> tuple[bool, int]:
    current_region_id = parse_region_id_from_url(page.url or "")
    current_region_name = await get_selected_region_name(page)

    if current_region_name == region_name and current_region_id:
        return True, current_region_id

    if not await open_region_modal(page):
        return False, 0

    clicked = False

    for selector in ("div[data-name='SpecialRegionsDesktop']", "div[data-name='PopularRegionsDesktop']"):
        if clicked:
            break
        try:
            target = page.locator(selector).get_by_text(region_name, exact=True).first
            if await target.count():
                try:
                    await target.scroll_into_view_if_needed(timeout=5000)
                except Exception:
                    pass
                try:
                    await target.click(timeout=5000)
                except Exception:
                    await target.click(timeout=5000, force=True)
                clicked = True
        except Exception:
            pass

    if not clicked:
        return False, 0

    for _ in range(30):
        await page.wait_for_timeout(500)
        region_id = parse_region_id_from_url(page.url or "")
        selected_name = await get_selected_region_name(page)
        if region_id and (selected_name == region_name or selected_name):
            if await has_agency_cards(page, 12_000):
                return True, region_id

    region_id = parse_region_id_from_url(page.url or "")
    if region_id and await has_agency_cards(page, 12_000):
        return True, region_id

    return False, 0


async def choose_targets_via_ui(ctx) -> List[str]:
    page = await new_stealth_page(ctx)
    try:
        warm_url = build_catalog_url(4630, 1)
        log.info("Открываю страницу для выбора региона: %s", warm_url)

        await page.goto(warm_url, wait_until="domcontentloaded", timeout=45_000)
        await page.wait_for_timeout(random.uniform(2500, 4500))

        region_names = await get_all_region_names(page)

        if not region_names:
            log.warning("Список регионов из UI не получен, сохраняю debug страницы")
            await debug_agencies_page(page)

        return region_names
    finally:
        await page.close()


async def resolve_region_names_via_ui(ctx, region_names: List[str]) -> List[int]:
    resolved: List[int] = []
    page = await new_stealth_page(ctx)
    try:
        if not await ensure_agencies_catalog_open(page):
            return []

        for region_name in region_names:
            ok, region_id = await select_region(page, region_name)
            if ok and region_id:
                resolved.append(region_id)
                log.info("Регион выбран через UI: %s -> regionId=%s", region_name, region_id)
            else:
                log.warning("Не удалось выбрать регион через UI: %s", region_name)

        return dedupe_region_ids(resolved)
    finally:
        await page.close()


async def debug_agencies_page(page) -> None:
    try:
        current_url = page.url
    except Exception:
        current_url = ""
    try:
        title = await page.title()
    except Exception:
        title = ""

    try:
        hrefs = await page.evaluate(
            """
            () => Array.from(document.querySelectorAll('a[href]'))
              .slice(0, 60)
              .map(a => a.getAttribute('href') || '')
              .filter(Boolean)
            """
        )
    except Exception:
        hrefs = []

    log.warning("Не найдены карточки агентств")
    log.warning("Debug URL: %s", current_url)
    log.warning("Debug title: %s", title)
    log.warning("Первые href: %s", hrefs[:30])

    html_file = OUTPUT_DIR / "debug_agencies_page.html"
    screenshot_file = OUTPUT_DIR / "debug_agencies_page.png"
    try:
        html_file.write_text(await page.content(), encoding="utf-8")
        log.warning("HTML сохранен: %s", html_file)
    except Exception as exc:
        log.warning("Не удалось сохранить HTML: %s", exc)
        debug_log.exception("Не удалось сохранить debug HTML | path=%s", html_file)
    try:
        await page.screenshot(path=str(screenshot_file), full_page=True)
        log.warning("Screenshot сохранен: %s", screenshot_file)
    except Exception as exc:
        log.warning("Не удалось сохранить screenshot: %s", exc)
        debug_log.exception("Не удалось сохранить debug screenshot | path=%s", screenshot_file)


async def get_last_page(page) -> int:
    state = await get_agents_listing_state(page)
    listing = ((state.get("agencies") or {}).get("listing") or {})
    total = listing.get("total") or 0
    limit = listing.get("limit") or 0
    if total and limit:
        return max(1, math.ceil(int(total) / int(limit)))

    try:
        items = page.locator("[data-name='PaginationWrapper'] [data-name='PaginationItem'] span")
        count = await items.count()
        numbers = []
        for i in range(count):
            text = normalize_ws(await items.nth(i).inner_text())
            if text.isdigit():
                numbers.append(int(text))
        return max(numbers) if numbers else 1
    except Exception:
        return 1


async def get_total_agencies(page, last_page: int) -> int:
    state = await get_agents_listing_state(page)
    listing = ((state.get("agencies") or {}).get("listing") or {})
    total = listing.get("total")
    try:
        total_int = int(total)
    except (TypeError, ValueError):
        total_int = 0
    return total_int if total_int > 0 else max(1, int(last_page or 1)) * 10


async def get_region_name(page, region_id: int) -> str:
    state = await get_agents_listing_state(page)
    geo = (((state.get("geo") or {}).get("locationsMap") or {}).get("agencies") or {})
    return normalize_ws(geo.get("displayName") or f"region_{region_id}")


async def extract_agency_cards(page) -> List[Dict]:
    state = await get_agents_listing_state(page)
    listing = ((state.get("agencies") or {}).get("listing") or {})
    items = listing.get("items") or []
    geo = (((state.get("geo") or {}).get("locationsMap") or {}).get("agencies") or {})
    region_name = normalize_ws(geo.get("displayName") or "")

    out: List[Dict] = []
    seen = set()

    if items:
        for item in items:
            agency_url = agency_url_from_id(item.get("cianUserId"))
            if not agency_url or agency_url in seen:
                continue
            seen.add(agency_url)
            row = {
                "agency_url": agency_url,
                "name": normalize_ws(item.get("name") or ""),
                "objects": normalize_offers_count(item.get("offersCount")),
                "rating": normalize_ws(item.get("userTrustLevelName") or ""),
                "region": region_name,
                "cian": normalize_cian_term(item.get("age") or ""),
                "card_details": normalize_ws(item.get("services") or ""),
            }
            out.append(row)
            log.info("Добавлена ссылка агентства: %s | %s", agency_url, row["name"])

        return out

    cards = page.locator('[data-name="AgencyCard"]')
    count = await cards.count()
    for i in range(count):
        card = cards.nth(i)
        text = normalize_ws(await text_or_empty(card))
        agency_url = ""
        name = ""

        try:
            href = await card.locator('a[href*="/company/"]').first.get_attribute("href")
            agency_url = agency_url_from_href(href or "")
        except Exception:
            pass

        if not agency_url:
            try:
                async with page.expect_popup(timeout=5000) as popup_info:
                    await card.click(timeout=5000)
                popup = await popup_info.value
                await popup.wait_for_load_state("domcontentloaded", timeout=30_000)
                agency_url = agency_url_from_href(popup.url)
                await popup.close()
            except Exception:
                agency_url = ""

        if not agency_url or agency_url in seen:
            continue

        lines = [normalize_ws(line) for line in re.split(r"\n+", text) if normalize_ws(line)]
        name = lines[0] if lines else ""
        seen.add(agency_url)
        row = {
            "agency_url": agency_url,
            "name": name,
            "objects": extract_objects_from_text(text),
            "rating": "Суперагент" if "Суперагент" in text else "",
            "region": region_name,
            "cian": extract_cian_term_from_text(text),
            "card_details": text,
        }
        out.append(row)
        log.info("Добавлена ссылка агентства: %s | %s", agency_url, row["name"])

    if out:
        return out

    try:
        links = await page.evaluate(
            """
            () => Array.from(document.querySelectorAll('a[href]')).map(a => ({
              href: a.getAttribute('href') || '',
              text: (a.innerText || a.textContent || '').trim().replace(/\\s+/g, ' ')
            })).filter(x => /\\/company\\/\\d+/.test(x.href) && !/\\/agents\\//.test(x.href))
            """
        )
    except Exception:
        links = []

    for link in links:
        agency_url = agency_url_from_href(link.get("href") or "")
        if not agency_url or agency_url in seen:
            continue
        seen.add(agency_url)
        row = {
            "agency_url": agency_url,
            "name": normalize_ws(link.get("text") or ""),
            "objects": "",
            "rating": "",
            "region": region_name,
            "cian": extract_cian_term_from_text(link.get("text") or ""),
            "card_details": "",
        }
        out.append(row)
        log.info("Добавлена ссылка агентства: %s | %s", agency_url, row["name"])

    return out


async def click_show_agency_phone(page) -> Dict:
    found_count = 0
    clicked_count = 0
    tel_count = 0
    clicked_keys = set()

    try:
        contacts = page.locator('[data-name="Contacts"]')
        contacts_count = await contacts.count()
        debug_step("Поиск кнопок Показать в Contacts", contacts_found=contacts_count)
        for contact_idx in range(contacts_count):
            rows = contacts.nth(contact_idx).locator('[data-name="AsideRow"]')
            row_count = await rows.count()
            for row_idx in range(row_count):
                row = rows.nth(row_idx)
                try:
                    row_text = normalize_ws(await row.inner_text(timeout=800))
                    if "Телефон:" not in row_text:
                        continue

                    show_buttons = row.locator('button:has-text("Показать"), a:has-text("Показать"), [role="button"]:has-text("Показать"), span:has-text("Показать"), div:has-text("Показать")')
                    button_count = await show_buttons.count()
                    found_count += button_count
                    for i in range(button_count):
                        if clicked_count >= 15:
                            break
                        locator = show_buttons.nth(i)
                        try:
                            box = await locator.bounding_box(timeout=500)
                            locator_text = normalize_ws(await locator.inner_text(timeout=300))
                            click_key = (
                                round((box or {}).get("x", 0)),
                                round((box or {}).get("y", 0)),
                                round((box or {}).get("width", 0)),
                                round((box or {}).get("height", 0)),
                                locator_text,
                            )
                        except Exception:
                            click_key = (contact_idx, row_idx, i)
                        if click_key in clicked_keys:
                            continue
                        clicked_keys.add(click_key)

                        try:
                            await locator.scroll_into_view_if_needed(timeout=1000)
                        except Exception:
                            pass
                        try:
                            await locator.click(timeout=1200)
                        except Exception:
                            await locator.click(timeout=1500, force=True)
                        clicked_count += 1
                        await page.wait_for_timeout(random.uniform(250, 450))

                        tel_count = await row.locator('a[href^="tel:"]').count()
                except Exception:
                    continue
            if clicked_count >= 15:
                break
    except Exception as exc:
        debug_log.exception("Ошибка при клике Показать в Contacts")
    contacts_data = await read_contacts_block(page)
    phones_saved = len(split_multiline_values(contacts_data.get("phone", "")))
    debug_step("Кнопки Показать обработаны", buttons_found=found_count, buttons_clicked=clicked_count, tel_links=tel_count, phones_saved=phones_saved)
    return {"buttons_found": found_count, "buttons_clicked": clicked_count, "tel_links": tel_count, "phones_saved": phones_saved}


async def read_contacts_block(page) -> Dict:
    phones: List[str] = []
    sites: List[str] = []
    region = ""
    total_rows = 0
    total_tel_links = 0

    try:
        contacts = page.locator('[data-name="Contacts"]')
        contacts_count = await contacts.count()
        debug_step("Проверка блока Contacts", contacts_found=contacts_count)
        for contact_idx in range(contacts_count):
            rows = contacts.nth(contact_idx).locator('[data-name="AsideRow"]')
            row_count = await rows.count()
            total_rows += row_count
            debug_step("Строки AsideRow в Contacts", contact_index=contact_idx + 1, aside_rows=row_count)
            for i in range(row_count):
                row = rows.nth(i)
                text = normalize_ws(await row.inner_text())
                low = text.lower()

                if "телефон:" in low:
                    tel_links = row.locator('a[href^="tel:"]')
                    tel_count = await tel_links.count()
                    total_tel_links += tel_count
                    debug_step("Телефонная строка Contacts", row_index=i + 1, tel_links=tel_count)
                    for j in range(tel_count):
                        href = await tel_links.nth(j).get_attribute("href") or ""
                        if not href.startswith("tel:"):
                            continue
                        phone = clean_phone(href.replace("tel:", "", 1).split("?", 1)[0])
                        if "XX" in phone.upper():
                            continue
                        append_unique(phones, phone)
                    for phone in split_multiline_values(extract_phones_from_text(text)):
                        if "XX" in phone.upper():
                            continue
                        append_unique(phones, phone)

                if "сайт:" in low:
                    value = normalize_ws(re.sub(r"(?i)^.*?сайт:\s*", "", text, count=1))
                    value = normalize_site_value(value)
                    if is_valid_agency_site(value):
                        append_unique_site(sites, value)

                if "регион работы:" in low and not region:
                    region = normalize_ws(re.sub(r"(?i)^.*?регион работы:\s*", "", text, count=1))
    except Exception as exc:
        debug_log.exception("Ошибка чтения Contacts")

    debug_step(
        "Итог чтения Contacts",
        aside_rows=total_rows,
        tel_links=total_tel_links,
        phones_found=len(phones),
        sites_found=len(sites),
        region_found=bool(region),
    )
    return {
        "phone": "\n".join(phones),
        "site": "\n".join(sites),
        "region": region,
    }


async def parse_agency_profile_page(page, url: str) -> Dict:
    data: Dict = {
        "agency_url": agency_url_from_href(url) or url,
        "name": "",
        "phone": "",
        "site": "",
        "region": "",
        "objects": "",
        "rating": "",
        "cian": "",
        "status": "Частично",
    }

    debug_step("Начало обработки профиля", url=url)
    for attempt in range(RETRY_COUNT):
        try:
            log.info("Открываю профиль агентства: %s", url)
            debug_step("Открытие профиля", url=url, attempt=attempt + 1)
            await page.goto(url, wait_until="domcontentloaded", timeout=45_000)
            await page.wait_for_timeout(random.uniform(1000, 1800))
            debug_step("Профиль открылся", url=url, attempt=attempt + 1)
            break
        except Exception as exc:
            log.warning("Профиль не открылся (%s/%s): %s | %s", attempt + 1, RETRY_COUNT, url, exc)
            debug_log.exception("Профиль не открылся | url=%s | attempt=%s", url, attempt + 1)
            if attempt + 1 == RETRY_COUNT:
                data["status"] = "Частично"
                debug_step("Профиль не открылся после всех попыток", url=url, status=data["status"])
                return data
            await page.wait_for_timeout(random.uniform(1500, 3000))

    try:
        await page.wait_for_selector('[data-name="AboutCompany"], [data-name="CompanyName"], [data-name="Contacts"]', timeout=15_000)
    except Exception:
        pass

    try:
        body = await page.evaluate("document.body ? document.body.innerText : ''")
    except Exception:
        body = ""

    data["name"] = await text_or_empty(page.locator('[data-name="CompanyName"]'))

    # 1. Сначала читаем текст страницы ДО клика
    try:
        body = await page.evaluate("document.body ? document.body.innerText : ''")
    except Exception:
        body = ""

    # 2. Пробуем раскрыть телефон
    click_stats = await click_show_agency_phone(page)
    debug_step("Раскрытие телефона завершено", url=url, buttons_found=click_stats.get("buttons_found"), buttons_clicked=click_stats.get("buttons_clicked"))

    # 3. Перечитываем Contacts коротко: дополнительные попытки нужны только после клика без телефона.
    contact_read_attempts = 3 if click_stats.get("buttons_clicked", 0) else 1
    for attempt in range(contact_read_attempts):
        contacts_data = await read_contacts_block(page)

        # Контакты: телефон, сайт и регион только из блока Contacts
        phones = split_multiline_values(data["phone"])
        for phone in split_multiline_values(contacts_data.get("phone", "")):
            if "XX" not in phone.upper():
                append_unique(phones, phone)
        data["phone"] = "\n".join(phones)

        sites = split_multiline_values(data["site"])
        for value in split_multiline_values(contacts_data.get("site", "")):
            value = normalize_site_value(value)
            if is_valid_agency_site(value):
                append_unique_site(sites, value)
        data["site"] = "\n".join(sites)

        if not data["region"] and contacts_data.get("region"):
            data["region"] = contacts_data["region"]

        # Если уже нашли основные поля — выходим раньше
        phone_full = data["phone"] and "XX" not in data["phone"].upper()
        if phone_full and data["region"]:
            break

        if phone_full or attempt + 1 >= contact_read_attempts:
            break

        await page.wait_for_timeout(500)

    try:
        headers = page.locator('[data-name="OfferGroupHeader"]')
        count = await headers.count()
        total = 0
        parts = []
        for i in range(count):
            text = normalize_ws(await headers.nth(i).inner_text())
            match = re.search(r"Смотреть все\s+(\d[\d\s]*)\s+предлож", text, re.I)
            if match:
                value = int(re.sub(r"\D", "", match.group(1)))
                total += value
                parts.append(match.group(0))
        if total:
            data["objects"] = f"{total} предложений"
        elif parts:
            data["objects"] = "; ".join(parts)
    except Exception:
        pass

    if body:
        if not data["objects"]:
            data["objects"] = extract_objects_from_text(body)
        if not data["cian"]:
            data["cian"] = extract_cian_term_from_text(body)

    if not data["name"]:
        title = ""
        try:
            title = await page.title()
        except Exception:
            pass
        data["name"] = normalize_ws(re.sub(r"\s*-\s*агентство.*$", "", title, flags=re.I))

    for key, value in list(data.items()):
        if isinstance(value, str):
            if key == "phone":
                data[key] = join_unique(split_multiline_values(value))
            elif key == "site":
                sites: List[str] = []
                for site in split_multiline_values(value):
                    append_unique_site(sites, site)
                data[key] = "\n".join(sites)
            elif key == "status":
                data[key] = normalize_status(value)
            else:
                data[key] = normalize_ws(value)

    phone_clean = data.get("phone", "")
    phone_is_full = bool(phone_clean) and "XX" not in phone_clean.upper()
    if phone_is_full and data["site"] and data["region"]:
        data["status"] = "Успешно"
    elif profile_has_enrichment(data):
        data["status"] = "Частично"
    else:
        data["status"] = "Частично"

    log.info(
        "Профиль обработан: name=%s | phone=%s | site=%s | region=%s | status=%s",
        data["name"] or "без названия",
        data["phone"] or "",
        data["site"] or "",
        data["region"] or "",
        data["status"] or "",
    )
    debug_step(
        "Итог профиля",
        url=url,
        name=data["name"],
        phone=data["phone"],
        site=data["site"],
        region=data["region"],
        status=data["status"],
    )
    return data


def merge_card_and_profile(card: Dict, profile: Dict) -> Dict:
    merged = dict(profile)
    merged["agency_url"] = profile.get("agency_url") or card.get("agency_url") or ""
    merged["name"] = profile.get("name") or card.get("name") or ""
    merged["objects"] = profile.get("objects") or card.get("objects") or ""
    merged["rating"] = profile.get("rating") or card.get("rating") or ""
    merged["cian"] = profile.get("cian") or card.get("cian") or ""
    merged["region"] = profile.get("region") or card.get("region") or ""
    phone_clean = merged.get("phone", "")
    phone_is_full = bool(phone_clean) and "XX" not in phone_clean.upper()
    merged["status"] = normalize_status(profile.get("status") or ("Успешно" if phone_is_full and merged.get("site") and merged.get("region") else "Частично"))
    return merged


def catalog_card_to_result(card: Dict) -> Dict:
    return {
        "agency_url": card.get("agency_url", ""),
        "name": card.get("name", ""),
        "phone": "",
        "site": "",
        "region": card.get("region", ""),
        "objects": card.get("objects", ""),
        "rating": card.get("rating", ""),
        "cian": card.get("cian", ""),
        "status": "Частично",
    }


async def enrich_profiles(
    ctx,
    cards: List[Dict],
    profile_tabs: int,
    region_name: str,
    progress_state: Optional[Dict] = None,
) -> List[Dict]:
    profile_tabs = max(1, min(int(profile_tabs or 1), 3))
    if not cards:
        return []

    queue: asyncio.Queue = asyncio.Queue()
    for idx, card in enumerate(cards):
        await queue.put((idx, card))

    results: List[Optional[Dict]] = [None] * len(cards)
    progress_lock = asyncio.Lock()
    processed_count = 0

    async def worker(worker_id: int) -> None:
        nonlocal processed_count
        page = await new_stealth_page(ctx)
        try:
            while True:
                try:
                    idx, card = queue.get_nowait()
                except asyncio.QueueEmpty:
                    break

                url = card.get("agency_url", "")
                if not url:
                    results[idx] = catalog_card_to_result(card)
                    results[idx]["status"] = "Частично"
                    async with progress_lock:
                        processed_count += 1
                        update_region_progress_stats(progress_state, results[idx], processed_increment=1)
                        log_region_profile_progress(progress_state)
                    queue.task_done()
                    continue

                row_result: Optional[Dict] = None
                try:
                    debug_step("Начало обработки профиля воркером", worker=worker_id, index=idx + 1, total=len(cards), url=url)
                    profile = await parse_agency_profile_page(page, url)
                    row_result = merge_card_and_profile(card, profile)
                    results[idx] = row_result
                    log.info("Вкладка %s: профиль %s/%s обработан", worker_id, idx + 1, len(cards))
                except asyncio.CancelledError as exc:
                    log.warning("Вкладка %s: профиль отменен, не заменяю прогресс карточкой каталога: %s | %s", worker_id, url, exc)
                    raise
                except Exception as exc:
                    log.warning("Вкладка %s: ошибка профиля, не заменяю прогресс карточкой каталога: %s | %s", worker_id, url, exc)
                    debug_log.exception("Ошибка обработки профиля воркером | worker=%s | url=%s", worker_id, url)
                    try:
                        await page.close()
                    except Exception:
                        pass
                    page = await new_stealth_page(ctx)
                finally:
                    async with progress_lock:
                        processed_count += 1
                        update_region_progress_stats(progress_state, row_result, processed_increment=1)
                        log_region_profile_progress(progress_state)

                    queue.task_done()
                    await asyncio.sleep(random.uniform(*PROFILE_DELAY))
        finally:
            try:
                await page.close()
            except Exception:
                pass

    tasks = [asyncio.create_task(worker(i + 1)) for i in range(min(profile_tabs, len(cards)))]
    try:
        await asyncio.gather(*tasks, return_exceptions=True)
    except (asyncio.CancelledError, KeyboardInterrupt):
        partial_results = [row for row in results if row]
        if partial_results:
            log_emergency_progress_save(partial_results, int((progress_state or {}).get("total_agencies", len(partial_results))))
            save_progress(partial_results, region_name)
        raise

    final_results = [row for row in results if row]
    return final_results


async def collect_agencies(
    ctx,
    region_id: int,
    pages: int,
    limit: int,
    catalog_only: bool = False,
    profile_tabs: int = 1,
) -> tuple[List[Dict], str]:

    cards: List[Dict] = []
    results: List[Dict] = []
    seen_urls = set()
    profiles_started = False
    region_name = f"region_{region_id}"
    region_started_at = time.monotonic()
    total_agencies = 0
    progress_state: Optional[Dict] = None
    debug_step("Старт сбора региона", region=region_name, region_id=region_id, catalog_only=catalog_only)

    catalog_page = await new_stealth_page(ctx)
    try:
        first_url = build_catalog_url(region_id, 1)
        log.info("Открываю каталог: %s", first_url)
        debug_step("Открытие страницы каталога", url=first_url, region=region_name, region_id=region_id, page=1)
        opened = False
        for attempt in range(RETRY_COUNT):
            try:
                await catalog_page.goto(first_url, wait_until="domcontentloaded", timeout=45_000)
                await catalog_page.wait_for_timeout(random.uniform(1800, 3000) if catalog_only else random.uniform(1200, 2200))
                opened = True
                break
            except asyncio.CancelledError as exc:
                log.warning("Открытие первой страницы каталога отменено (%s/%s): %s", attempt + 1, RETRY_COUNT, exc)
                await catalog_page.wait_for_timeout(random.uniform(2500, 4500))
            except Exception as exc:
                log.warning("Не открылась первая страница каталога (%s/%s): %s", attempt + 1, RETRY_COUNT, exc)
                debug_log.exception("Не открылась первая страница каталога | url=%s | attempt=%s", first_url, attempt + 1)
                await catalog_page.wait_for_timeout(random.uniform(2500, 4500))

        if not opened:
            debug_step("Страница каталога не открылась", url=first_url, region=region_name, region_id=region_id)
            return [], region_name

        if not await has_agency_cards(catalog_page):
            debug_step("Карточки агентств не найдены", url=first_url, region=region_name, region_id=region_id)
            await debug_agencies_page(catalog_page)
            return [], region_name

        region_name = await get_region_name(catalog_page, region_id)
        debug_step("Выбранный регион", region=region_name, region_id=region_id)
        last_page = await get_last_page(catalog_page)
        if pages > 0:
            last_page = min(last_page, pages)
        total_agencies = await get_total_agencies(catalog_page, last_page)
        progress_state = {
            "region_name": region_name,
            "total_agencies": total_agencies,
            "processed_profiles": 0,
            "phones": 0,
            "sites": 0,
            "successful": 0,
            "partial": 0,
            "started_at": region_started_at,
        }
        log.info("Регион: %s | страниц: 1/%s | агентств всего: %s", region_name, last_page, total_agencies)

        for page_num in range(1, last_page + 1):
            url = build_catalog_url(region_id, page_num)
            log.info("Открываю URL каталога: %s", url)
            debug_step("Открытие страницы каталога", url=url, region=region_name, region_id=region_id, page=page_num)
            if page_num > 1:
                page_opened = False
                for attempt in range(RETRY_COUNT):
                    try:
                        await catalog_page.goto(url, wait_until="domcontentloaded", timeout=45_000)
                        await catalog_page.wait_for_timeout(
                            random.uniform(1200, 2200) if catalog_only else random.uniform(1000, 2000)
                        )
                        page_opened = True
                        break
                    except asyncio.CancelledError as exc:
                        log.warning("Открытие каталога отменено %s (%s/%s): %s", url, attempt + 1, RETRY_COUNT, exc)
                        await catalog_page.wait_for_timeout(random.uniform(2500, 4500))
                    except Exception as exc:
                        log.warning("Не открылся каталог %s (%s/%s): %s", url, attempt + 1, RETRY_COUNT, exc)
                        debug_log.exception("Не открылся каталог | url=%s | attempt=%s", url, attempt + 1)
                        await catalog_page.wait_for_timeout(random.uniform(2500, 4500))
                if not page_opened:
                    log.warning("Пропускаю страницу каталога после ошибок: %s", url)
                    debug_step("Пропуск страницы каталога после ошибок", url=url, page=page_num)
                    continue

            try:
                if not await has_agency_cards(catalog_page):
                    await debug_agencies_page(catalog_page)
                    continue
            except asyncio.CancelledError as exc:
                log.warning("Проверка карточек отменена на странице %s: %s", page_num, exc)
                continue
            except Exception as exc:
                log.warning("Ошибка проверки карточек на странице %s: %s", page_num, exc)
                debug_log.exception("Ошибка проверки карточек | page=%s | url=%s", page_num, url)
                continue

            try:
                page_cards = await extract_agency_cards(catalog_page)
            except asyncio.CancelledError as exc:
                log.warning("Сбор карточек отменен на странице %s: %s", page_num, exc)
                page_cards = []
            except Exception as exc:
                log.warning("Ошибка сбора карточек на странице %s: %s", page_num, exc)
                debug_log.exception("Ошибка сбора карточек | page=%s | url=%s", page_num, url)
                page_cards = []
            log.info("Карточек найдено на странице %s: %s", page_num, len(page_cards))
            debug_step("Карточек найдено", page=page_num, cards_found=len(page_cards), url=url)

            new_cards: List[Dict] = []

            for card in page_cards:
                agency_url = card.get("agency_url", "")
                if not agency_url or agency_url in seen_urls:
                    continue

                seen_urls.add(agency_url)
                cards.append(card)
                new_cards.append(card)

                if limit and len(cards) >= limit:
                    break

            log.info(
                "Страница %s/%s | карточек на странице: %s | собрано карточек: %s/%s",
                page_num,
                last_page,
                len(page_cards),
                len(cards),
                total_agencies,
            )
            progress_file: Optional[Path] = None

            if catalog_only:
                progress_file = save_progress([catalog_card_to_result(card) for card in cards], region_name)
            else:
                if new_cards:
                    profiles_started = True
                    log.info(
                        "Обрабатываю профили со страницы %s: %s шт.",
                        page_num,
                        len(new_cards),
                    )

                    enriched_page_results = await enrich_profiles(
                        ctx=ctx,
                        cards=new_cards,
                        profile_tabs=profile_tabs,
                        region_name=region_name,
                        progress_state=progress_state,
                    )

                    results.extend(enriched_page_results)
                    progress_file = save_progress(results, region_name)
                elif results:
                    progress_file = save_progress(results, region_name)

            page_rows = results if not catalog_only else [catalog_card_to_result(card) for card in cards]
            page_phone_count, page_site_count = count_progress_contacts(page_rows)
            processed_profiles = int((progress_state or {}).get("processed_profiles", len(results)))
            if progress_file:
                log.info(
                    "Итог страницы %s/%s: обработано профилей %s/%s | телефоны %s | сайты %s | файл progress сохранён",
                    page_num,
                    last_page,
                    processed_profiles,
                    total_agencies,
                    page_phone_count,
                    page_site_count,
                )
            else:
                log.info(
                    "Итог страницы %s/%s: обработано профилей %s/%s | телефоны %s | сайты %s | файл progress не обновлён",
                    page_num,
                    last_page,
                    processed_profiles,
                    total_agencies,
                    page_phone_count,
                    page_site_count,
                )

            if limit and len(cards) >= limit:
                log.info("Достигнут limit=%s", limit)
                break

            delay = (0.4, 0.9) if catalog_only else CATALOG_DELAY
            await asyncio.sleep(random.uniform(*delay))

        if catalog_only:
            log.info("Режим catalog-only: профили агентств не открываются")
            return [catalog_card_to_result(card) for card in cards], region_name

        return results, region_name

    except asyncio.CancelledError:
        debug_step("collect_agencies отменен", region=region_name, region_id=region_id, processed_results=len(results), cards=len(cards))
        if results:
            log_emergency_progress_save(results, total_agencies)
            save_progress(results, region_name)
        elif cards and (catalog_only or not profiles_started):
            emergency_rows = [catalog_card_to_result(card) for card in cards]
            log_emergency_progress_save(emergency_rows, total_agencies)
            save_progress(emergency_rows, region_name)
        raise

    except KeyboardInterrupt:
        debug_step("collect_agencies остановлен KeyboardInterrupt", region=region_name, region_id=region_id, processed_results=len(results), cards=len(cards))
        if results:
            log_emergency_progress_save(results, total_agencies)
            save_progress(results, region_name)
        elif cards and (catalog_only or not profiles_started):
            emergency_rows = [catalog_card_to_result(card) for card in cards]
            log_emergency_progress_save(emergency_rows, total_agencies)
            save_progress(emergency_rows, region_name)
        raise

    except Exception as exc:
        debug_log.exception("Ошибка collect_agencies | region=%s | regionId=%s", region_name, region_id)
        if results:
            log_emergency_progress_save(results, total_agencies)
            save_progress(results, region_name)
        elif cards and (catalog_only or not profiles_started):
            emergency_rows = [catalog_card_to_result(card) for card in cards]
            log_emergency_progress_save(emergency_rows, total_agencies)
            save_progress(emergency_rows, region_name)
        raise

    finally:
        try:
            await catalog_page.close()
        except Exception:
            pass


def save_to_excel(data: List[Dict], region_name: str, filename: Optional[Path] = None) -> Path:
    tag = sanitize_tag(region_name)
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = OUTPUT_DIR / f"cian_agencies_{tag}_{timestamp}.xlsx"
    phone_count = sum(1 for row in data if normalize_ws(row.get("phone", "")))
    site_count = sum(1 for row in data if normalize_ws(row.get("site", "")))
    debug_step(
        "Сохранение Excel",
        path=filename,
        rows=len(data),
        rows_with_phone=phone_count,
        rows_with_site=site_count,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Агентства ЦИАН"

    headers = [
        "Ссылка",
        "Название",
        "На Циан",
        "Телефоны",
        "Сайт",
        "Регион",
        "Объекты",
        "Рейтинг",
        "Статус",
    ]

    hfont = Font(bold=True, color="000000", size=11)
    link_font = Font(color="1A56DB", underline="single", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = hfont
        cell.alignment = center
        cell.border = border

    widths = [46, 30, 18, 24, 28, 36, 18, 18, 16]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    fill = PatternFill()

    def line_count(value: str, width: int) -> int:
        value = str(value or "").replace("\r\n", "\n").replace("\r", "\n")
        if not value:
            return 1
        total = 0
        for part in value.split("\n"):
            total += max(1, (len(part.strip()) + max(width, 8) - 1) // max(width, 8))
        return total

    for row_idx, row in enumerate(data, 2):
        values = [
            row.get("agency_url", ""),
            row.get("name", ""),
            row.get("cian", ""),
            row.get("phone", ""),
            row.get("site", ""),
            row.get("region", ""),
            row.get("objects", ""),
            row.get("rating", ""),
            normalize_status(row.get("status", "")),
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = fill
            cell.border = border
            cell.alignment = left
            if col == 1 and value:
                cell.font = link_font
                cell.hyperlink = value

        max_lines = 1
        for col, value in enumerate(values, 1):
            max_lines = max(max_lines, line_count(value, widths[col - 1]))
        ws.row_dimensions[row_idx].height = min(240, 14 * max_lines + 4)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(data) + 1}"
    wb.save(filename)
    log.info("Excel сохранен: %s", filename)
    debug_step("Excel сохранен", path=filename, rows=len(data), rows_with_phone=phone_count, rows_with_site=site_count)
    return filename


def dedupe_region_ids(region_ids: List[int]) -> List[int]:
    out: List[int] = []
    seen = set()
    for region_id in region_ids:
        if region_id in seen:
            continue
        seen.add(region_id)
        out.append(region_id)
    return out


def progress_filename(region_name: str) -> Path:
    return OUTPUT_DIR / f"cian_agencies_{sanitize_tag(region_name)}_PROGRESS.xlsx"


def row_has_contacts(row: Dict) -> bool:
    return bool(normalize_ws(row.get("phone", "")) or normalize_ws(row.get("site", "")))


def count_progress_contacts(rows: List[Dict]) -> tuple[int, int]:
    phones = sum(1 for row in rows if normalize_ws(row.get("phone", "")))
    sites = sum(1 for row in rows if normalize_ws(row.get("site", "")))
    return phones, sites


def count_progress_statuses(rows: List[Dict]) -> tuple[int, int]:
    successful = sum(1 for row in rows if normalize_status(row.get("status", "")) == normalize_status("ok"))
    partial = sum(1 for row in rows if normalize_status(row.get("status", "")) == normalize_status("partial"))
    return successful, partial


def update_region_progress_stats(stats: Optional[Dict], row: Optional[Dict] = None, processed_increment: int = 1) -> None:
    if stats is None:
        return

    stats["processed_profiles"] = int(stats.get("processed_profiles", 0)) + processed_increment
    if not row:
        return

    if normalize_ws(row.get("phone", "")):
        stats["phones"] = int(stats.get("phones", 0)) + 1
    if normalize_ws(row.get("site", "")):
        stats["sites"] = int(stats.get("sites", 0)) + 1

    status = normalize_status(row.get("status", ""))
    if status == normalize_status("ok"):
        stats["successful"] = int(stats.get("successful", 0)) + 1
    elif status == normalize_status("partial"):
        stats["partial"] = int(stats.get("partial", 0)) + 1


def log_region_profile_progress(stats: Optional[Dict]) -> None:
    if stats is None:
        return

    processed = int(stats.get("processed_profiles", 0))
    total = max(int(stats.get("total_agencies", 0) or 0), processed)
    remaining = max(total - processed, 0)
    log.info(
        "Прогресс региона %s: обработано %s/%s | с телефонами: %s | с сайтами: %s | успешно: %s | частично: %s | осталось: %s",
        stats.get("region_name", ""),
        processed,
        total,
        int(stats.get("phones", 0)),
        int(stats.get("sites", 0)),
        int(stats.get("successful", 0)),
        int(stats.get("partial", 0)),
        remaining,
    )

    elapsed_minutes = max((time.monotonic() - float(stats.get("started_at", time.monotonic()))) / 60, 1 / 60)
    speed = processed / elapsed_minutes if processed else 0.0
    eta_minutes = math.ceil(remaining / speed) if speed > 0 and remaining > 0 else 0
    log.info("Скорость: %.1f проф/мин | осталось примерно: %s мин", speed, eta_minutes)


def log_emergency_progress_save(rows: List[Dict], total_agencies: int) -> None:
    phone_count, site_count = count_progress_contacts(rows)
    total = max(int(total_agencies or 0), len(rows))
    log.info(
        "Аварийное сохранение: обработано %s/%s | с телефонами: %s | с сайтами: %s",
        len(rows),
        total,
        phone_count,
        site_count,
    )


def progress_row_key(row: Dict) -> str:
    return normalize_ws(row.get("agency_url", ""))


def merge_progress_row(existing: Dict, incoming: Dict) -> Dict:
    if row_has_contacts(existing) and not row_has_contacts(incoming):
        return existing

    merged = dict(existing)
    for key, value in incoming.items():
        value = normalize_ws(value) if isinstance(value, str) else value
        if value:
            merged[key] = value

    return merged


def merge_progress_rows(existing_rows: List[Dict], incoming_rows: List[Dict]) -> List[Dict]:
    merged_rows: List[Dict] = []
    positions: Dict[str, int] = {}

    for row in existing_rows:
        key = progress_row_key(row)
        if key:
            positions[key] = len(merged_rows)
        merged_rows.append(dict(row))

    for row in incoming_rows:
        key = progress_row_key(row)
        if key and key in positions:
            idx = positions[key]
            merged_rows[idx] = merge_progress_row(merged_rows[idx], row)
            continue
        if key:
            positions[key] = len(merged_rows)
        merged_rows.append(dict(row))

    return merged_rows


def load_progress_rows(filename: Path) -> List[Dict]:
    if not filename.exists():
        return []

    header_map = {
        "Ссылка": "agency_url",
        "Название": "name",
        "Название агентства": "name",
        "На Циан": "cian",
        "ЦИАН": "cian",
        "Телефоны": "phone",
        "Телефон": "phone",
        "Сайт": "site",
        "Регион": "region",
        "Объекты": "objects",
        "Рейтинг": "rating",
        "Статус": "status",
    }
    try:
        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        ws = wb.active
        headers = [header_map.get(normalize_ws(str(cell.value or "")), "") for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows: List[Dict] = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            row: Dict = {}
            for idx, key in enumerate(headers):
                if not key:
                    continue
                row[key] = normalize_ws(str(values[idx] or "") if idx < len(values) else "")
            if progress_row_key(row):
                rows.append(row)
        wb.close()
        return rows
    except Exception as exc:
        log.warning("Не удалось прочитать текущий progress-файл %s: %s", filename, exc)
        debug_log.exception("Не удалось прочитать текущий progress-файл | path=%s", filename)
        return []


def save_progress(data: List[Dict], region_name: str) -> Optional[Path]:
    try:
        filename = progress_filename(region_name)
        existing_rows = load_progress_rows(filename)
        rows = merge_progress_rows(existing_rows, data)
        if not rows:
            return None
        phone_count, site_count = count_progress_contacts(rows)
        log.info(
            "Сохраняю прогресс: строк=%s | с телефонами=%s | с сайтами=%s | файл=%s",
            len(rows),
            phone_count,
            site_count,
            filename,
        )
        try:
            save_to_excel(rows, region_name, filename=filename)
        except PermissionError:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = OUTPUT_DIR / f"cian_agencies_{sanitize_tag(region_name)}_PROGRESS_{timestamp}.xlsx"
            log.info(
                "Сохраняю прогресс в новый файл: строк=%s | с телефонами=%s | с сайтами=%s | файл=%s",
                len(rows),
                phone_count,
                site_count,
                filename,
            )
            save_to_excel(rows, region_name, filename=filename)
        log.info("Промежуточное сохранение: %s строк -> %s", len(rows), filename)
        return filename
    except Exception as exc:
        log.warning("Не удалось выполнить промежуточное сохранение для %s: %s", region_name, exc)
        debug_log.exception("Не удалось выполнить промежуточное сохранение | region=%s", region_name)
        return None


def parse_region_ids_arg(value: str) -> List[int]:
    region_ids: List[int] = []
    for part in (value or "").split(","):
        part = part.strip()
        if not part:
            continue
        if not part.isdigit():
            raise ValueError(f"Некорректный regionId: {part}")
        region_ids.append(int(part))
    return dedupe_region_ids(region_ids)


def print_start_banner() -> None:
    print("=" * 60)
    print("  Парсер агентств ЦИАН")
    print("=" * 60)
    if STEALTH_AVAILABLE:
        print("\n  ✅ playwright-stealth активен\n")
    else:
        print("\n  ⚠️  playwright-stealth не установлен или недоступен\n")


def print_available_regions(region_names: List[str]) -> None:
    print("\nДоступные регионы:")
    for idx, name in enumerate(region_names, 1):
        print(f"[{idx:3}] {name}")


def choose_mode(region_names: List[str]) -> tuple[str, List[str]]:
    print("=" * 60)
    print("  Выберите режим:")
    print("=" * 60)
    print("  [0]  Все регионы")
    print("  [1]  Один регион")
    print("  [2]  Диапазон регионов (с N-го по M-й)")
    print("=" * 60)
    while True:
        mode = input("  Введите номер: ").strip()

        if mode == "0":
            return "Все регионы", region_names

        if mode == "1":
            print_available_regions(region_names)
            while True:
                raw = input("\nВведите номер региона: ").strip()
                if raw.isdigit():
                    idx = int(raw)
                    if 1 <= idx <= len(region_names):
                        return "Один регион", [region_names[idx - 1]]
                print("Неверный выбор.")

        if mode == "2":
            print_available_regions(region_names)
            while True:
                raw_from = input(f"\nС какого номера (1-{len(region_names)}): ").strip()
                raw_to = input(f"По какой номер   (1-{len(region_names)}, Enter = до конца): ").strip()
                if raw_from.isdigit():
                    start = int(raw_from)
                    end = int(raw_to) if raw_to.isdigit() else len(region_names)
                    if 1 <= start <= end <= len(region_names):
                        return f"С {start} по {end} регион", region_names[start - 1 : end]
                print("Неверный диапазон.")

        print("Неверный выбор.")


def should_use_interactive(args: argparse.Namespace) -> bool:
    explicit_region_mode = (
        args.region_id is not None
        or bool(args.region_ids)
        or args.all_regions
    )
    return args.interactive or not explicit_region_mode


def resolve_cli_region_ids(args: argparse.Namespace) -> tuple[str, List[int]]:
    if args.all_regions:
        return "Все регионы из списка REGIONS", list(REGIONS.keys())

    if args.region_ids:
        return "Несколько регионов по ID", parse_region_ids_arg(args.region_ids)

    if args.region_id is not None:
        return "Один регион", [args.region_id]

    return "Не выбран", []


def print_run_summary(
    mode_name: str,
    region_count: int,
    pages: int,
    limit: int,
    headless: bool,
    catalog_only: bool,
    profile_tabs: int,
) -> None:
    print("=" * 60)
    print(f"Режим         : {mode_name}")
    print(f"Регионов      : {region_count}")
    print(f"Страниц       : {'все' if pages == 0 else pages}")
    print(f"Лимит         : {limit if limit else 'нет'}")
    print(f"Headless      : {'да' if headless else 'нет'}")
    print(f"Catalog-only  : {'да' if catalog_only else 'нет'}")
    print(f"Profile tabs  : {profile_tabs}")
    print("=" * 60)
    input("Нажмите Enter для старта...")


async def run_regions(
    ctx,
    region_ids: List[int],
    pages: int,
    limit: int,
    catalog_only: bool,
    profile_tabs: int,
) -> List[Dict]:
    all_results: List[Dict] = []

    for idx, region_id in enumerate(region_ids, 1):
        configured_name = REGIONS.get(region_id, f"region_{region_id}")
        log.info("Старт региона %s/%s: %s (regionId=%s)", idx, len(region_ids), configured_name, region_id)
        debug_step("Выбранный регион для запуска", index=idx, total=len(region_ids), region=configured_name, region_id=region_id)

        results, region_name = await collect_agencies(
            ctx=ctx,
            region_id=region_id,
            pages=pages,
            limit=limit,
            catalog_only=catalog_only,
            profile_tabs=profile_tabs,
        )

        if results:
            output = save_to_excel(results, region_name or configured_name)
            phone_count, site_count = count_progress_contacts(results)
            successful, partial = count_progress_statuses(results)
            log.info(
                "Регион завершён: %s | всего строк: %s | с телефонами: %s | с сайтами: %s | успешно: %s | частично: %s | файл: %s",
                region_name or configured_name,
                len(results),
                phone_count,
                site_count,
                successful,
                partial,
                output,
            )
            print(f"{region_name or configured_name}: собрано агентств: {len(results)}")
            print(f"Excel: {output}")
            all_results.extend(results)
        else:
            log.warning("Данные агентств не собраны для региона %s (regionId=%s)", configured_name, region_id)

        if idx < len(region_ids):
            pause = random.uniform(*REGION_PAUSE)
            log.info("Пауза перед следующим регионом: %.1f сек.", pause)
            await asyncio.sleep(pause)

    if len(region_ids) > 1 and all_results:
        output_all = save_to_excel(all_results, "ALL")
        print(f"Общий Excel: {output_all}")

    phone_count, site_count = count_progress_contacts(all_results)
    successful = sum(1 for row in all_results if normalize_status(row.get("status", "")) == "Успешно")
    partial = sum(1 for row in all_results if normalize_status(row.get("status", "")) == "Частично")
    debug_step(
        "Итог парсинга",
        total_rows=len(all_results),
        rows_with_phone=phone_count,
        rows_with_site=site_count,
        successful=successful,
        partial=partial,
    )
    return all_results


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Парсер агентств ЦИАН")
    parser.add_argument("--region-id", type=int, default=None, help="ID региона ЦИАН")
    parser.add_argument("--region-ids", default="", help="Несколько regionId через запятую, например 4630,1,2")
    parser.add_argument("--all-regions", action="store_true", help="Обойти все регионы из локального списка REGIONS")
    parser.add_argument("--interactive", action="store_true", help="Выбрать регионы через консольное меню")
    parser.add_argument("--pages", type=int, default=1, help="Количество страниц, 0 означает все страницы")
    parser.add_argument("--headless", action="store_true", help="Запустить браузер без окна")
    parser.add_argument("--limit", type=int, default=0, help="Ограничение количества агентств для теста")
    parser.add_argument("--catalog-only", action="store_true", help="Собрать только карточки каталога без открытия профилей")
    parser.add_argument("--profile-tabs", type=int, default=1, help="Количество вкладок для профилей: 1-3")
    return parser.parse_args(argv)


async def run_app(args: argparse.Namespace) -> None:
    from playwright.async_api import async_playwright

    args.profile_tabs = max(1, min(int(args.profile_tabs or 1), 3))
    debug_step(
        "Старт программы",
        pages=args.pages,
        limit=args.limit,
        headless=args.headless,
        catalog_only=args.catalog_only,
        profile_tabs=args.profile_tabs,
        debug_log=DEBUG_LOG_FILE,
    )

    if not STEALTH_AVAILABLE:
        log.warning("playwright-stealth не установлен или недоступен, продолжаю без stealth")
    else:
        log.info("playwright-stealth активен")

    interactive_mode = should_use_interactive(args)
    run_pages = args.pages
    run_limit = args.limit
    run_catalog_only = args.catalog_only
    run_profile_tabs = args.profile_tabs

    async with async_playwright() as pw:
        browser, ctx = await make_browser_context(pw, args.headless)
        try:
            if interactive_mode:
                run_pages = 0
                run_limit = 0
                run_catalog_only = False
                run_profile_tabs = 1

                print_start_banner()
                print("  Получаю список регионов из UI каталога агентств...\n")
                region_names = await choose_targets_via_ui(ctx)
                if not region_names:
                    log.warning("Не удалось получить список регионов из UI ЦИАН. Использую fallback-список названий регионов.")
                    region_names = FALLBACK_REGION_NAMES

                mode_name, selected_region_names = choose_mode(region_names)
                if not selected_region_names:
                    log.warning("Не выбраны регионы для парсинга")
                    return

                print_run_summary(
                    mode_name=mode_name,
                    region_count=len(selected_region_names),
                    pages=run_pages,
                    limit=run_limit,
                    headless=args.headless,
                    catalog_only=run_catalog_only,
                    profile_tabs=run_profile_tabs,
                )

                region_ids = await resolve_region_names_via_ui(ctx, selected_region_names)
            else:
                try:
                    mode_name, region_ids = resolve_cli_region_ids(args)
                except ValueError as exc:
                    log.error("%s", exc)
                    return

                if not region_ids:
                    log.warning("Не выбраны регионы для парсинга")
                    return

            if not region_ids:
                log.warning("Не удалось определить regionId для выбранных регионов")
                return

            log.info("Выбраны регионы: %s", ", ".join(str(region_id) for region_id in region_ids))
            debug_step("Список выбранных regionId", region_ids=", ".join(str(region_id) for region_id in region_ids))
            await run_regions(
                ctx=ctx,
                region_ids=region_ids,
                pages=run_pages,
                limit=run_limit,
                catalog_only=run_catalog_only,
                profile_tabs=run_profile_tabs,
            )
        except Exception as exc:
            debug_log.exception("Ошибка верхнего уровня run_app")
            raise
        finally:
            await browser.close()


def main() -> None:
    args = parse_args(sys.argv[1:])
    try:
        asyncio.run(run_app(args))
    except KeyboardInterrupt:
        debug_step("Парсер остановлен KeyboardInterrupt")
        raise
    except Exception as exc:
        debug_log.exception("Аварийное завершение программы")
        raise


if __name__ == "__main__":
    main()
